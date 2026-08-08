# -*- coding: utf-8 -*-
import json
from collections import Counter, defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation

F=json.load(open("analise_final.json")); OCC=json.load(open("ocorrencias.json"))
inst=F["instrumentos"]; gT=F["graus_todos"]; gV=F["graus_vigentes"]; mod=F["forca3_modalidade"]; pds=F["pd_status"]
vig_de={x["arquivo"]:x["vigencia_juridica"] for x in inst}; rev_de={x["arquivo"]:x["revisao_status"] for x in inst}

MUNIS=["Belford Roxo","Duque de Caxias","Japeri","Mesquita","Nilopolis","Nova Iguaçu","Queimados","São João de Meriti"]
TIPOS=["CTM","PGV","Plano Diretor","Plano Mobilidade","Plano Saneamento","Plano Diretor de Tecnologia"]
LEIS=["Estatuto da Cidade (Lei 10.257/2001)","INDE - Decreto 6.666/2008","Decreto 12.402/2025 (atualiza INDE)",
 "Politica Nacional de Desenvolvimento Urbano (PNDU)","Estatuto da Metropole (Lei 13.089/2015)",
 "Lei 13.683/2018 (altera Estatuto Metropole)","Portaria 3.242/2022 (CTM)","IN RFB 2.030/2021 (CIB)","Decreto 11.208/2022 (SINTER)"]
LEIS_CURTA=["Estatuto Cidade","INDE","Dec 12.402/25","PNDU","Est. Metropole","Lei 13.683/18","Portaria 3.242 (CTM)","CIB","SINTER"]
ANO_LEI=[2001,2008,2025,2001,2015,2018,2022,2021,2022]

FONT="Calibri"
H=Font(name=FONT,bold=True,color="FFFFFF",size=11); HF=PatternFill("solid",fgColor="1F4E79")
SUBF=PatternFill("solid",fgColor="D6E4F0"); TITLE=Font(name=FONT,bold=True,size=15,color="1F4E79")
NB=Font(name=FONT,size=10); CTR=Alignment(horizontal="center",vertical="center",wrap_text=True)
WRAP=Alignment(wrap_text=True,vertical="top")
thin=Side(style="thin",color="BFBFBF"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
GRAU_COLOR={"Consolidado":"C6EFCE","Parcial":"FFEB9C","Incipiente":"FCE4D6"}
def graucol(g): return next((v for k,v in GRAU_COLOR.items() if g and g.startswith(k)),"F2F2F2")
def catcol(c):
    if c.startswith("Explícita (forte)"):return "C6EFCE"
    if c.startswith("Explícita"):return "D9EAD3"
    if c.startswith("Implícita"):return "FFEB9C"
    if c.startswith("Prática"):return "FCE4D6"
    return "F2F2F2"
def vigcol(v): return "D9D9D9" if v.startswith("Substituído") else "C6EFCE"
def revcol(r):
    if r.startswith("Norma recente"):return "C6EFCE"
    if "mantido por leis" in r.lower():return "FFEB9C"
    if "SEM revisão" in r or "defasado" in r.lower():return "FCE4D6"
    if r.startswith("Base antiga"):return "FFF2CC"
    return "F2F2F2"
CAM_COLOR={"DIRETO":"C6EFCE","GENERICO":"FFEB9C","PRATICA":"FCE4D6","PERIFERICO":"F2F2F2"}
GMAP={"Consolidado":3,"Parcial":2,"Incipiente":1}

def hdr(ws,headers,row=1,h=30):
    for j,t in enumerate(headers,1):
        c=ws.cell(row=row,column=j,value=t); c.font=H; c.fill=HF; c.alignment=CTR; c.border=BORDER
    ws.row_dimensions[row].height=h
def setw(ws,W):
    for j,w in enumerate(W,1): ws.column_dimensions[get_column_letter(j)].width=w

wb=Workbook(); wb.remove(wb.active)

# ---- folha de dados dos graficos ----
dz=wb.create_sheet("Dados_Graficos")
CATS4=["Explícita","Implícita","Prática","Ausente"]
ccount=Counter(x["categoria_v2"].split(" (")[0] for x in inst)
dz["A1"]="cat"; dz["B1"]="n"
for i,c in enumerate(CATS4,2): dz.cell(row=i,column=1,value=c); dz.cell(row=i,column=2,value=ccount.get(c,0))
dz["D1"]="muni"; dz["E1"]="idade_PD"
for i,m in enumerate(MUNIS,2):
    dz.cell(row=i,column=4,value=m); dz.cell(row=i,column=5,value=pds[m]["idade"] if pds[m]["idade"] is not None else 0)
dz["H1"]="muni"; dz["I1"]="norm"
for i,m in enumerate(MUNIS,2):
    dz.cell(row=i,column=8,value=m); dz.cell(row=i,column=9,value=gT[m]["norm"])
dz.sheet_state="hidden"

# ============ LEIA-ME ============
ws=wb.create_sheet("Leia-me"); setw(ws,[120])
L=[("Institucionalização da Geoinformação — vigência, revisão e conteúdo",TITLE),
 ("Tese · Seção 3.2.1 · Baixada Fluminense (RJ) · 32 instrumentos (nenhum excluído)",NB),("",NB),
 (f"Operantes (em vigor): {F['n_operantes']} | Substituídos: {F['n_substituidos']} | Planos Diretores defasados (>10 anos sem revisão): {F['n_pd_defasados']}",Font(name=FONT,bold=True,size=11)),("",NB),
 ("DOIS EIXOS DISTINTOS (importante)",Font(name=FONT,bold=True,size=12)),
 ("1) VIGÊNCIA JURÍDICA — a norma está em vigor? Quase todas estão; só os Planos Diretores antigos de Queimados e",NB),
 ("   São João foram revogados por versão posterior (Substituídos).",NB),
 ("2) REVISÃO / ATUALIDADE — houve revisão completa dentro do prazo? Leis complementares emendam pontualmente, mas",NB),
 ("   NÃO equivalem a uma revisão do Plano Diretor. A revisão decenal (art. 40 §3º do Estatuto da Cidade) continua devida.",NB),
 ("Um Plano Diretor pode estar VIGENTE e, ao mesmo tempo, DEFASADO (sem revisão completa) — são coisas diferentes.",Font(name=FONT,bold=True,size=10)),("",NB),
 ("DEFASAGEM ≠ AUSÊNCIA DE GEOINFORMAÇÃO",Font(name=FONT,bold=True,size=12)),
 ("Estar defasado não retira o conteúdo técnico: o texto pode mencionar geoprocessamento ou demonstrar a prática na",NB),
 ("elaboração de mapas. Ex.: o PD de Duque de Caxias é defasado (20 anos) mas tem conteúdo 'Explícita (forte)'.",NB),
 ("Por isso o conteúdo de geoinformação é analisado em TODOS os documentos; a defasagem é um flag de conformidade à parte.",NB),("",NB),
 ("ABAS",Font(name=FONT,bold=True,size=12)),
 ("- Dashboard / Situacao_Instrumentos / PD_Defasagem / Sintese / Robustez_Grau / IIG_Dois_Eixos /",NB),
 ("  Modalidade_Forca3 / Leis_Federais / Sensibilidade / Catalogo_Ocorrencias / Verificacao_Pares / Termos.",NB),("",NB),
 ("Toda classificação automática é ponto de partida, sujeita a verificação por pares posterior.",Font(name=FONT,bold=True,size=11,color="C00000")),
 ("A coluna 'Revisão/atualidade' traz um padrão automático que VOCÊ deve confirmar (ex.: quais PDs têm leis complementares).",Font(name=FONT,size=10,color="C00000")),
]
for i,(t,f) in enumerate(L,1):
    c=ws.cell(row=i,column=1,value=t); c.font=f; c.alignment=Alignment(wrap_text=True,vertical="top")

# ============ SITUACAO_INSTRUMENTOS ============
ws=wb.create_sheet("Situacao_Instrumentos")
hdr(ws,["Município","Instrumento","Arquivo","Ano-base","Última atualiz. (proxy)","Ano efetivo","Vigência jurídica","Revisão / atualidade (confirmar)","PD defasado?","Observação (editável)","Conferido p/ pares (S/N)"])
setw(ws,[15,19,30,11,16,11,22,38,11,38,15])
r=2
for x in sorted(inst,key=lambda a:(a["municipio"],TIPOS.index(a["instrumento"]) if a["instrumento"] in TIPOS else 9)):
    vals=[x["municipio"],x["instrumento"],x["arquivo"],x["ano_base"],x["ano_atualizacao"],x["ano_efetivo"],
          x["vigencia_juridica"],x["revisao_status"],"Sim" if x["pd_defasado"] else "",x["observacao"],""]
    for j,v in enumerate(vals,1):
        c=ws.cell(row=r,column=j,value=v); c.font=NB; c.border=BORDER; c.alignment=(WRAP if j in(3,7,8,10) else CTR)
    ws.cell(row=r,column=7).fill=PatternFill("solid",fgColor=vigcol(x["vigencia_juridica"]))
    ws.cell(row=r,column=8).fill=PatternFill("solid",fgColor=revcol(x["revisao_status"]))
    if x["pd_defasado"]: ws.cell(row=r,column=9).fill=PatternFill("solid",fgColor="F4B6B6")
    r+=1
ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:K{r-1}"
dv=DataValidation(type="list",formula1='"S,N"',allow_blank=True); ws.add_data_validation(dv); dv.add(f"K2:K{r-1}")

# ============ PD_DEFASAGEM ============
ws=wb.create_sheet("PD_Defasagem")
ws.cell(row=1,column=1,value="Planos Diretores: vigência × revisão decenal (art. 40 §3º da Lei 10.257/2001)").font=TITLE
ws.merge_cells("A1:F1")
ws.cell(row=2,column=1,value="A lei exige revisão completa ao menos a cada 10 anos. Leis complementares emendam, mas não substituem a revisão. Defasagem é flag legal, não de conteúdo: o PD pode estar defasado e ainda conter geoinformação.").font=Font(name=FONT,size=9,italic=True)
ws.merge_cells("A2:F2"); ws.row_dimensions[2].height=28
hdr(ws,["Município","Ano do PD em vigor","Idade (anos)","Conformidade (revisão decenal)","Conteúdo de geoinformação do PD","Observação"],row=4)
setw(ws,[16,16,11,26,28,34])
r=5
for m in MUNIS:
    s=pds[m]; ano=s["ano"]; idade=s["idade"]; defas=s["defasado"]; cat=s.get("categoria")
    leg="DEFASADO (> 10 anos, sem revisão)" if defas else ("Dentro do prazo" if ano else "PD não localizado")
    obs="Revisão decenal vencida — leis complementares não suprem" if defas else ("" if ano else "Sem Plano Diretor no acervo")
    vals=[m,ano if ano else "—",idade if idade is not None else "—",leg,cat or "—",obs]
    for j,v in enumerate(vals,1):
        c=ws.cell(row=r,column=j,value=v); c.font=NB; c.border=BORDER; c.alignment=(WRAP if j in(4,5,6) else CTR)
    ws.cell(row=r,column=4).fill=PatternFill("solid",fgColor="F4B6B6" if defas else ("C6EFCE" if ano else "F2F2F2"))
    if cat: ws.cell(row=r,column=5).fill=PatternFill("solid",fgColor=catcol(cat))
    r+=1
ws.cell(row=r+1,column=1,value=f"{F['n_pd_defasados']} de {len(MUNIS)} municípios com Plano Diretor defasado. Observe que a defasagem é independente do conteúdo: há PDs defasados com geoinformação 'Explícita' (ex.: Duque de Caxias).").font=Font(name=FONT,bold=True,color="C00000")
ws.merge_cells(start_row=r+1,start_column=1,end_row=r+1,end_column=6); ws.cell(row=r+1,column=1).alignment=Alignment(wrap_text=True)

# ============ SINTESE ============
ws=wb.create_sheet("Sintese")
cols=["Município","Instrumento","Arquivo","Ano efetivo","Vigência","Revisão/atualidade","Categoria geoinfo (c/ piso)","Eixo precisão","Eixo prática",
 "Nº menções diretas","Nº ocorrências","Soma peso","Força máx","Nº temas","Leis (aplicáveis)","Boilerplate?"]
hdr(ws,cols)
setw(ws,[14,18,28,9,12,30,22,11,14,9,10,9,8,7,13,26])
r=2
for x in sorted(inst,key=lambda a:(TIPOS.index(a["instrumento"]) if a["instrumento"] in TIPOS else 9,a["municipio"])):
    vals=[x["municipio"],x["instrumento"],x["arquivo"],x["ano_efetivo"],x["vigencia_juridica"].split(" (")[0],x["revisao_status"],
     x["categoria_v2"],x["eixo_precisao"],x["eixo_pratica"],x["n_ocorr_diretas"],x["n_ocorrencias"],x["soma_peso"],
     x["forca_max"],x["n_temas"],x["leis_citadas_aplic"],x["leis_boilerplate"]]
    for j,v in enumerate(vals,1):
        c=ws.cell(row=r,column=j,value=v); c.font=NB; c.border=BORDER; c.alignment=(WRAP if j in(3,6,7,16) else CTR)
    ws.cell(row=r,column=5).fill=PatternFill("solid",fgColor=vigcol(x["vigencia_juridica"]))
    ws.cell(row=r,column=6).fill=PatternFill("solid",fgColor=revcol(x["revisao_status"]))
    ws.cell(row=r,column=7).fill=PatternFill("solid",fgColor=catcol(x["categoria_v2"]))
    r+=1
ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:{get_column_letter(len(cols))}{r-1}"

# ============ ROBUSTEZ_GRAU ============
ws=wb.create_sheet("Robustez_Grau")
ws.cell(row=1,column=1,value="Robustez do grau — intensidade bruta vs normalizada por página").font=TITLE
ws.merge_cells("A1:G1")
hdr(ws,["Município","Grau (rubrica)","Soma peso","Páginas","Intensidade (peso/1000 pág.)","Rank bruto","Rank normaliz."],row=3)
setw(ws,[18,18,14,10,18,11,15])
rb=sorted(MUNIS,key=lambda m:-(gT[m]["soma"] or 0)); rn=sorted(MUNIS,key=lambda m:-(gT[m]["norm"] or 0))
rank_b={m:i+1 for i,m in enumerate(rb)}; rank_n={m:i+1 for i,m in enumerate(rn)}
r=4
for m in MUNIS:
    a=gT[m]
    vals=[m,a["grau"],a["soma"],a["pgs"],a["norm"],rank_b[m],rank_n[m]]
    for j,v in enumerate(vals,1):
        c=ws.cell(row=r,column=j,value=v); c.font=NB; c.border=BORDER; c.alignment=CTR
    ws.cell(row=r,column=2).fill=PatternFill("solid",fgColor=graucol(a["grau"]))
    if rank_b[m]!=rank_n[m]: ws.cell(row=r,column=7).fill=PatternFill("solid",fgColor="FFF2CC")
    r+=1
ws.cell(row=r+1,column=1,value="Controlando o tamanho do documento (intensidade por página), o ranking muda: municípios menores sobem. O grau bruto reflete em parte a extensão do texto. Células amarelas = posição muda entre os rankings.").font=Font(name=FONT,size=9,italic=True)
ws.merge_cells(start_row=r+1,start_column=1,end_row=r+2,end_column=7); ws.cell(row=r+1,column=1).alignment=Alignment(wrap_text=True)
ch=BarChart(); ch.type="col"; ch.title="Intensidade normalizada (peso/1000 pág.)"; ch.height=7; ch.width=16; ch.legend=None
ch.add_data(Reference(dz,min_col=9,min_row=2,max_row=1+len(MUNIS))); ch.set_categories(Reference(dz,min_col=8,min_row=2,max_row=1+len(MUNIS)))
ch.series[0].graphicalProperties.solidFill="264478"; ch.y_axis.majorGridlines=None
ws.add_chart(ch,f"A{r+4}")

# ============ IIG_DOIS_EIXOS ============
ws=wb.create_sheet("IIG_Dois_Eixos")
ws.cell(row=1,column=1,value="IIG de dois eixos: precisão terminológica × prática espacial (todos os instrumentos)").font=TITLE
ws.merge_cells("A1:C1")
ws.cell(row=2,column=1,value="Eixo vertical = como o documento NOMEIA a geoinformação; horizontal = se há FERRAMENTAS/coleta de dado espacial.").font=Font(name=FONT,size=9,italic=True)
ws.merge_cells("A2:C2")
hdr(ws,["Precisão \\ Prática","Com prática espacial","Sem prática"],row=4)
setw(ws,[20,40,40])
grid=defaultdict(list)
for x in inst: grid[(x["eixo_precisao"],x["eixo_pratica"])].append(f"{x['municipio']} ({x['instrumento']})")
r=5
for pl in ["Direta","Genérica","Nenhuma"]:
    ws.cell(row=r,column=1,value=pl).font=Font(name=FONT,bold=True,size=10); ws.cell(row=r,column=1).fill=SUBF; ws.cell(row=r,column=1).border=BORDER
    for j,pr in enumerate(["Com prática espacial","Sem prática"],2):
        items=grid.get((pl,pr),[])
        c=ws.cell(row=r,column=j,value="\n".join(items) if items else "—")
        c.font=Font(name=FONT,size=9); c.alignment=WRAP; c.border=BORDER
        c.fill=PatternFill("solid",fgColor="C6EFCE" if pl=="Direta" else ("FFEB9C" if pl=="Genérica" else "F2F2F2"))
    ws.row_dimensions[r].height=80; r+=1

# ============ MODALIDADE_FORCA3 ============
ws=wb.create_sheet("Modalidade_Forca3")
ws.cell(row=1,column=1,value="Modalidade das menções de força técnica 3 — PRÉ-PREENCHIDO p/ validação por pares").font=TITLE
ws.merge_cells("A1:J1")
hdr(ws,["Município","Instrumento","Ano efet.","Vigência","Termo","Pág.","Modalidade (automática)","Modalidade VALIDADA","Trecho","Contexto"],row=3)
setw(ws,[15,19,9,14,20,6,26,22,20,62])
r=4
for o in sorted(mod,key=lambda a:(a["municipio"],a["instrumento"])):
    vals=[o["municipio"],o["instrumento"],o.get("ano_efetivo"),(o.get("situacao") or "").split(" (")[0],o["termo"],o["pagina"],
          o["modalidade_auto"],"",o["trecho"],o["contexto"]]
    for j,v in enumerate(vals,1):
        c=ws.cell(row=r,column=j,value=v); c.font=Font(name=FONT,size=9); c.alignment=(WRAP if j in(9,10) else CTR); c.border=BORDER
    mc={"Estruturante (institui/cria)":"C6EFCE","Aspiracional (prevê/pretende)":"FFEB9C","Descritiva (uso/método)":"FCE4D6","Mencao simples":"F2F2F2"}
    ws.cell(row=r,column=7).fill=PatternFill("solid",fgColor=mc.get(o["modalidade_auto"],"FFFFFF"))
    r+=1
ws.freeze_panes="A4"; ws.auto_filter.ref=f"A3:J{r-1}"
dv2=DataValidation(type="list",formula1='"Estruturante,Aspiracional,Descritiva,Mencao simples,Falso positivo"',allow_blank=True)
ws.add_data_validation(dv2); dv2.add(f"H4:H{r-1}")

# ============ LEIS_FEDERAIS ============
ws=wb.create_sheet("Leis_Federais")
hdr(ws,["Município","Instrumento","Ano efet."]+LEIS_CURTA)
setw(ws,[15,19,9]+[12]*9)
r=2
for x in sorted(inst,key=lambda a:(a["municipio"],TIPOS.index(a["instrumento"]) if a["instrumento"] in TIPOS else 9)):
    ws.cell(row=r,column=1,value=x["municipio"]).font=NB; ws.cell(row=r,column=2,value=x["instrumento"]).font=NB
    ws.cell(row=r,column=3,value=x["ano_efetivo"]).font=NB; ws.cell(row=r,column=3).alignment=CTR
    ano=x["ano_efetivo"] or 0
    for j,(lei,aly) in enumerate(zip(LEIS,ANO_LEI),4):
        has=lei in x["leis_federais"]
        if aly>ano and ano:
            c=ws.cell(row=r,column=j,value="n/a"); c.fill=PatternFill("solid",fgColor="E0E0E0"); c.font=Font(name=FONT,size=8,italic=True,color="888888")
        elif has:
            c=ws.cell(row=r,column=j,value="X"); c.fill=PatternFill("solid",fgColor="C6EFCE"); c.font=Font(name=FONT,bold=True,color="1F4E79")
        else: c=ws.cell(row=r,column=j,value="")
        c.alignment=CTR; c.border=BORDER
    for j in (1,2): ws.cell(row=r,column=j).border=BORDER
    r+=1
ws.freeze_panes="D2"; ws.auto_filter.ref=f"A1:L{r-1}"
ws.cell(row=r+1,column=1,value="'n/a' = norma federal posterior ao ano do documento (controle temporal). Verde = citação detectada.").font=Font(name=FONT,size=9,italic=True)
ws.merge_cells(start_row=r+1,start_column=1,end_row=r+1,end_column=12); ws.cell(row=r+1,column=1).alignment=Alignment(wrap_text=True)

# ============ SENSIBILIDADE ============
ws=wb.create_sheet("Sensibilidade")
ws.cell(row=1,column=1,value="Análise de sensibilidade — efeito de cada decisão metodológica").font=TITLE
ws.merge_cells("A1:D1")
hdr(ws,["Decisão metodológica","O que muda","Efeito observado","Implicação"],row=3)
setw(ws,[26,30,40,34])
n_aus=sum(1 for x in inst if "Ausente" in x["categoria_v2"])
n_boil=sum(1 for x in inst if "boilerplate" in x["leis_boilerplate"])
rows=[
 ("Vigência × revisão (2 eixos)","Separa norma em vigor de revisão completa","Quase todos vigentes; 2 substituídos; leis complementares não contam como revisão","Evita confundir 'em vigor' com 'atualizado'."),
 ("Regra decenal do Plano Diretor","Aplica o art. 40 §3º do Estatuto da Cidade",f"{F['n_pd_defasados']} municípios com PD defasado (>10 anos)","Defasagem vira evidência de descumprimento legal."),
 ("Defasagem ≠ conteúdo","Mantém a análise de geoinfo em docs defasados","PD de Duque de Caxias: defasado mas 'Explícita (forte)'","Conteúdo técnico é independente da revisão."),
 ("Piso 'Ausente'","Não conta cadastro/zoneamento isolado como institucionalização",f"{n_aus} instrumentos vão para 'Ausente'","Remove falsa institucionalização por termos periféricos."),
 ("Normalização por página","Divide a intensidade pelo tamanho do documento","Reordena o ranking; municípios menores sobem","Grau bruto é em parte prolixidade."),
 ("Modalidade das força-3","Distingue institui/cria de menção/método","0 estruturante, 2 aspiracionais, 14 descritivas, 65 menções (1º passe)","Sustenta 'presença textual ≠ força institucional'."),
 ("Boilerplate federal","Separa citação deliberada de texto de lei transcrito",f"{n_boil} instrumentos com citação federal marcada como herdada","Adesão federal pode estar superestimada."),
]
r=4
for a,b,c,d in rows:
    for j,v in enumerate([a,b,c,d],1):
        cc=ws.cell(row=r,column=j,value=v); cc.font=NB; cc.border=BORDER; cc.alignment=WRAP
    ws.cell(row=r,column=1).font=Font(name=FONT,bold=True,size=10); ws.row_dimensions[r].height=44; r+=1

# ============ CATALOGO ============
ws=wb.create_sheet("Catalogo_Ocorrencias")
cols=["ID","Vigência doc.","Município","Instrumento","Arquivo","Pág.","Origem","Termo-chave","Tema","Camada","Força","Peso",
      "Trecho","Contexto (≈300 car.)","Validado (S/N)","Relevante (S/N)","Notas do pesquisador"]
hdr(ws,cols,h=28)
setw(ws,[5,14,14,18,28,6,8,20,18,11,7,8,20,58,12,13,26])
r=2
for idx,o in enumerate(sorted(OCC,key=lambda a:(a["municipio"],-a["peso"])),1):
    vj=vig_de.get(o["arquivo"],"Vigente")
    vals=[idx,vj.split(" (")[0],o["municipio"],o["instrumento"],o["arquivo"],o["pagina"],o["origem"],o["termo"],o["tema"],
          o["camada"],o["forca"],o["peso"],o["trecho"],o["contexto"],"","",""]
    for j,v in enumerate(vals,1):
        c=ws.cell(row=r,column=j,value=v); c.font=Font(name=FONT,size=9); c.alignment=WRAP; c.border=BORDER
    ws.cell(row=r,column=2).fill=PatternFill("solid",fgColor=vigcol(vj))
    ws.cell(row=r,column=10).fill=PatternFill("solid",fgColor=CAM_COLOR.get(o["camada"],"FFFFFF")); ws.cell(row=r,column=10).alignment=CTR
    if o["forca"]==3: ws.cell(row=r,column=11).fill=PatternFill("solid",fgColor="C6EFCE")
    elif o["forca"]==2: ws.cell(row=r,column=11).fill=PatternFill("solid",fgColor="FFEB9C")
    ws.cell(row=r,column=11).alignment=CTR
    r+=1
ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:{get_column_letter(len(cols))}{r-1}"
dv3=DataValidation(type="list",formula1='"S,N"',allow_blank=True); ws.add_data_validation(dv3); dv3.add(f"O2:P{r-1}")

# ============ VERIFICACAO_PARES ============
ws=wb.create_sheet("Verificacao_Pares"); setw(ws,[4,40,40,30])
ws.cell(row=1,column=1,value="Protocolo de verificação por pares (2ª codificação independente)").font=TITLE
ws.merge_cells("A1:D1")
passos=[
 "1. Amostra: todas as força-3 (Modalidade_Forca3) + 20% aleatório das demais ocorrências (Catalogo).",
 "2. 2º codificador, sem ver a classificação automática, marca: o trecho trata de geoinformação? (S/N) e a modalidade.",
 "3. Confirmar, na aba Situacao_Instrumentos, a coluna 'Revisão/atualidade' (quais PDs têm leis complementares) e as datas.",
 "4. Comparar as duas codificações e calcular a concordância (kappa de Cohen).",
 "5. Resolver divergências por consenso; registrar regras novas nas observações.",
 "6. Recalcular categorias e grau após o consenso; versionar a planilha.",
]
r=3
for p in passos:
    ws.cell(row=r,column=1,value="•").font=Font(name=FONT,bold=True)
    cc=ws.cell(row=r,column=2,value=p); cc.font=NB; cc.alignment=WRAP; ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=4); r+=1
r+=1; ws.cell(row=r,column=1,value="Ficha de registro").font=Font(name=FONT,bold=True,size=12); r+=1
hdr(ws,["Item","Campo","Valor","Observações"],row=r); r+=1
for f in ["Codificador 1 (automático)","Codificador 2 (revisor humano)","Data da revisão","Nº de itens conferidos","Nº de divergências","Concordância (kappa)","Regras ajustadas","Versão final aprovada por"]:
    ws.cell(row=r,column=2,value=f).font=NB
    for cc in range(1,5): ws.cell(row=r,column=cc).border=BORDER
    r+=1

# ============ TERMOS ============
ws=wb.create_sheet("Termos")
hdr(ws,["Termo-chave","Camada","Força","Tema","Nº ocorrências","Nº documentos","Municípios"])
setw(ws,[28,12,8,26,15,13,12])
agg=defaultdict(lambda:{"n":0,"docs":set(),"munis":set(),"camada":"","forca":0,"tema":""})
for o in OCC:
    a=agg[o["termo"]]; a["n"]+=1; a["docs"].add(o["arquivo"]); a["munis"].add(o["municipio"]); a["camada"]=o["camada"]; a["forca"]=o["forca"]; a["tema"]=o["tema"]
r=2
for termo,a in sorted(agg.items(),key=lambda x:-x[1]["n"]):
    for j,v in enumerate([termo,a["camada"],a["forca"],a["tema"],a["n"],len(a["docs"]),len(a["munis"])],1):
        c=ws.cell(row=r,column=j,value=v); c.font=NB; c.border=BORDER; c.alignment=(WRAP if j in(1,4) else CTR)
    ws.cell(row=r,column=2).fill=PatternFill("solid",fgColor=CAM_COLOR.get(a["camada"],"FFFFFF")); r+=1
ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:G{r-1}"

# ============ DASHBOARD ============
ws=wb.create_sheet("Dashboard"); wb.move_sheet("Dashboard", -(len(wb.sheetnames)-1))
ws.sheet_view.showGridLines=False
ws.merge_cells("A1:N1"); ws["A1"]="PAINEL — Institucionalização da Geoinformação (32 instrumentos)"; ws["A1"].font=TITLE
ws.merge_cells("A2:N2"); ws["A2"]="Baixada Fluminense (RJ) · vigência × revisão × conteúdo · sujeito a verificação por pares"; ws["A2"].font=Font(name=FONT,size=10,italic=True,color="595959")
kpis=[("Instrumentos (total)",F["n_total"],"1F4E79"),
 ("Em vigor (operantes)",F["n_operantes"],"2E7D32"),
 ("Substituídos",F["n_substituidos"],"888888"),
 ("PD defasados (>10 anos)",F["n_pd_defasados"],"C0392B"),
 ("Explícita (forte)",sum(1 for x in inst if x["categoria_v2"].startswith("Explícita (forte)")),"2E7D32"),
 ("Citam INDE/SINTER/CIB",sum(1 for x in inst if any(k in x["leis_federais"] for k in ["INDE","SINTER","CIB"])),"C0392B")]
col=1;r0=4
for label,val,color in kpis:
    ws.merge_cells(start_row=r0,start_column=col,end_row=r0,end_column=col+1)
    ws.merge_cells(start_row=r0+1,start_column=col,end_row=r0+2,end_column=col+1)
    cl=ws.cell(row=r0,column=col,value=label); cl.font=Font(name=FONT,size=9,bold=True,color="FFFFFF"); cl.alignment=CTR; cl.fill=PatternFill("solid",fgColor=color)
    cv=ws.cell(row=r0+1,column=col,value=val); cv.font=Font(name=FONT,size=22,bold=True,color=color); cv.alignment=CTR; cv.fill=PatternFill("solid",fgColor="F2F6FB")
    for rr in (r0,r0+1,r0+2):
        for cc in (col,col+1): ws.cell(row=rr,column=cc).border=BORDER
    col+=2
for c in range(1,15): ws.column_dimensions[get_column_letter(c)].width=11
ch=BarChart(); ch.type="col"; ch.title="Instrumentos por categoria de geoinformação (com piso)"; ch.height=7.5; ch.width=12; ch.legend=None
ch.add_data(Reference(dz,min_col=2,min_row=2,max_row=1+len(CATS4))); ch.set_categories(Reference(dz,min_col=1,min_row=2,max_row=1+len(CATS4)))
ch.series[0].graphicalProperties.solidFill="5B9BD5"; ch.y_axis.majorGridlines=None
ws.add_chart(ch,"A8")
ch2=BarChart(); ch2.type="col"; ch2.title="Idade do Plano Diretor por município (limite legal = 10 anos)"; ch2.height=7.5; ch2.width=14; ch2.legend=None
ch2.add_data(Reference(dz,min_col=5,min_row=2,max_row=1+len(MUNIS))); ch2.set_categories(Reference(dz,min_col=4,min_row=2,max_row=1+len(MUNIS)))
ch2.series[0].graphicalProperties.solidFill="C0392B"; ch2.y_axis.majorGridlines=None
ws.add_chart(ch2,"H8")

wb.save("Analise_Institucionalizacao_Geoinformacao.xlsx")
print("OK. Abas:",wb.sheetnames)
print("Operantes:",F["n_operantes"],"Substituidos:",F["n_substituidos"],"PD defasados:",F["n_pd_defasados"])
