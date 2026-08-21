# -*- coding: utf-8 -*-
"""
Analise metodologica dos instrumentos urbanisticos (tese - secao 3.2.1)
Institucionalizacao da geoinformacao nos municipios.

- Le os parametros (vocabulario, forca tecnica, temas, rubrica) do mesmo
  arquivo usado na analise das LDOs/LOAs: Parametros_LDO_LOA.xlsx
- Extrai texto dos PDFs (PyMuPDF) e aplica OCR (Tesseract) nas paginas digitalizadas
- Normaliza (minusculas, sem acento), busca termos via regex, mede a ancoragem
  territorial pelo contexto, calcula peso = forca x ancoragem
- Classifica a institucionalizacao (Explicita / Implicita / Pratica / Ausente)
- Detecta mencoes a leis federais
- Exporta JSON + CSV (consolidado e ocorrencias) para montar a planilha

Uso:
    python3 analise_instrumentos.py --pasta "<.../Instrumentos>" \
        --parametros "<.../Parametros_LDO_LOA.xlsx>" --saida "<.../outputs>"
"""
import os, re, sys, json, glob, argparse, unicodedata, datetime, hashlib

# ----------------------------------------------------------------------------
# Dependencias opcionais (degrada com elegancia se faltar)
# ----------------------------------------------------------------------------
try:
    import fitz  # PyMuPDF
except Exception as e:
    print("ERRO: PyMuPDF (fitz) e obrigatorio.", e); sys.exit(1)
try:
    import openpyxl
except Exception:
    openpyxl = None
try:
    import pytesseract
    from PIL import Image
    import io
    import _ocr_env
    # Usa a MESMA logica de descoberta do Tesseract das outras etapas do pipeline
    # (variavel TESSERACT_CMD -> caminhos padrao do Windows -> PATH) e confirma que o
    # programa realmente executa, em vez de so' supor que existe.
    TEM_OCR, _OCR_INFO = _ocr_env.configurar_ocr()
    print(("[OCR] disponivel: " + _OCR_INFO) if TEM_OCR
          else ("[OCR] INDISPONIVEL: " + _OCR_INFO))
except Exception as _e:
    TEM_OCR = False
    print(f"[OCR] INDISPONIVEL: {_e}")
try:
    from rapidfuzz import fuzz
    TEM_FUZZY = True
except Exception:
    TEM_FUZZY = False

# ----------------------------------------------------------------------------
# Normalizacao
# ----------------------------------------------------------------------------
def normalizar(txt: str) -> str:
    if not txt:
        return ""
    txt = unicodedata.normalize("NFKD", txt)
    txt = "".join(c for c in txt if not unicodedata.combining(c))
    txt = txt.lower()
    txt = re.sub(r"-\n", "", txt)          # junta palavras quebradas por hifen
    txt = re.sub(r"\s*\n\s*", " ", txt)
    txt = re.sub(r"[ \t]+", " ", txt)
    return txt

# ----------------------------------------------------------------------------
# Carregamento dos parametros (do Excel da LDO) com fallback embutido
# ----------------------------------------------------------------------------
def normalizar_regex(pat: str) -> str:
    pat = unicodedata.normalize("NFKD", pat)
    pat = "".join(c for c in pat if not unicodedata.combining(c))
    return pat

def carregar_parametros(xlsx_path):
    P = {"termos": [], "tema": {}, "forca": {}, "territoriais": set(),
         "administrativas": set(), "falso_positivo": set(), "fuzzy": [],
         "rubrica": [], "geral": {}}
    if openpyxl and xlsx_path and os.path.exists(xlsx_path):
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)

        def linhas(aba):
            if aba not in wb.sheetnames:
                return
            ws = wb[aba]
            rows = list(ws.iter_rows(values_only=True))
            hdr_i = 0
            for i, r in enumerate(rows[:5]):
                cells = [str(c).strip().lower() if c is not None else "" for c in r]
                if any(h in cells for h in ("termo_chave","parametro","lista","nivel","termo_canonico")):
                    hdr_i = i; break
            hdr = [str(c).strip().lower() if c is not None else "" for c in rows[hdr_i]]
            for r in rows[hdr_i+1:]:
                if r is None or all(c is None for c in r):
                    continue
                yield dict(zip(hdr, r))

        for d in linhas("Termos_Busca"):
            tk = (d.get("termo_chave") or "").strip()
            pat = (d.get("padrao_regex") or "").strip()
            if tk and pat:
                P["termos"].append((tk, normalizar_regex(pat)))
        for d in linhas("Tema_Padrao"):
            tk = (d.get("termo_chave") or "").strip()
            if tk: P["tema"][tk] = (d.get("tema_padrao") or "").strip()
        for d in linhas("Forca_Tecnica"):
            tk = (d.get("termo_chave") or "").strip()
            if tk:
                try: P["forca"][tk] = int(d.get("forca_tecnica"))
                except Exception: pass
        for d in linhas("Vocabulario_Apoio"):
            lista = (d.get("lista") or "").strip().upper()
            val = (d.get("valor") or "").strip().lower()
            if not val: continue
            if lista == "PALAVRAS_TERRITORIAIS": P["territoriais"].add(val)
            elif lista == "PALAVRAS_ADMINISTRATIVAS": P["administrativas"].add(val)
            elif lista in ("FALSO_POSITIVO","FALSOS_POSITIVOS"): P["falso_positivo"].add(val)
        for d in linhas("Termos_Canon_Fuzzy"):
            t = (d.get("termo_canonico") or "").strip().lower()
            if t: P["fuzzy"].append(t)
        for d in linhas("Grau_Rubrica"):
            if d.get("nivel"): P["rubrica"].append(d)
        for d in linhas("Parametros_Gerais"):
            p = (d.get("parametro") or "").strip()
            if p: P["geral"][p] = d.get("valor")
    return P

# ----------------------------------------------------------------------------
# Camadas de institucionalizacao (mapeia termo_chave -> camada)
# ----------------------------------------------------------------------------
CAMADA = {
    "geoinformação": "DIRETO", "geoprocessamento": "DIRETO", "geotecnologia": "DIRETO",
    "sistema de informação geográfica": "DIRETO",
    "Infraestrutura Nacional de Dados Espaciais": "DIRETO",
    "banco de dados geográfico": "DIRETO", "cadastro multifinalitário": "DIRETO",
    "dados espaciais": "GENERICO", "cartografia": "GENERICO", "mapeamento": "GENERICO",
    "imagem de satélite": "PRATICA", "sensoriamento remoto": "PRATICA",
    "drone": "PRATICA", "levantamento de imagens": "PRATICA",
    "cadastro imobiliário": "PERIFERICO", "cadastro fiscal": "PERIFERICO",
    "planta genérica de valores": "PERIFERICO", "regularização fundiária": "PERIFERICO",
    "zoneamento": "PERIFERICO", "uso e ocupação do solo": "PERIFERICO",
    "cadastro social": "PERIFERICO",
}

PRATICA_EXTRA = {
    "georreferenciamento": r"\bgeorreferencia(?:mento|d[oa]s?|r)\b",
    "ortofoto": r"\bortofoto(?:carta)?s?\b",
    "aerolevantamento": r"\baerolevantamento\b|\baerofotogrametri[ac]o?\b",
    "base cartográfica": r"\bbases? cartografica?s?\b",
    "dados territoriais": r"\bdados territoriais\b",
    "geodésico": r"\bgeodesic[oa]s?\b|\bgeodesia\b",
}

LEIS_FEDERAIS = {
    "Estatuto da Cidade (Lei 10.257/2001)":
        r"\bestatuto da cidade\b|\blei\s*(?:n[oº\.]*\s*)?10\.?257\b",
    "INDE - Decreto 6.666/2008":
        r"\bdecreto\s*(?:n[oº\.]*\s*)?6\.?666\b|infraestrutura nacional de dados espaciais|\binde\b",
    "Decreto 12.402/2025 (atualiza INDE)":
        r"\bdecreto\s*(?:n[oº\.]*\s*)?12\.?402\b",
    "Politica Nacional de Desenvolvimento Urbano (PNDU)":
        r"politica nacional de desenvolvimento urbano|\bpndu\b",
    "Estatuto da Metropole (Lei 13.089/2015)":
        r"\bestatuto da metropole\b|\blei\s*(?:n[oº\.]*\s*)?13\.?089\b|\bpdui\b",
    "Lei 13.683/2018 (altera Estatuto Metropole)":
        r"\blei\s*(?:n[oº\.]*\s*)?13\.?683\b",
    "Portaria 3.242/2022 (CTM)":
        r"\bportaria\s*(?:n[oº\.]*\s*)?3\.?242\b",
    "IN RFB 2.030/2021 (CIB)":
        r"\b2\.?030\b.{0,20}\brfb\b|cadastro imobiliario brasileiro|\bcib\b",
    "Decreto 11.208/2022 (SINTER)":
        r"\bdecreto\s*(?:n[oº\.]*\s*)?11\.?208\b|\bsinter\b|sistema nacional de gestao de informacoes territoriais",
}

MUNICIPIOS = ["Belford Roxo","Duque de Caxias","Japeri","Mesquita","Nilopolis",
              "Nova Iguaçu","Queimados","São João de Meriti"]
def parse_nome(fn):
    base = os.path.basename(fn).replace(".pdf","")
    partes = base.split(" - ", 1)
    tipo = partes[0].strip()
    muni = "?"
    nb = normalizar(base)
    for m in MUNICIPIOS:
        if normalizar(m) in nb:
            muni = m; break
    obs = []
    if "antig" in base.lower(): obs.append("versao antiga")
    if "atual" in base.lower(): obs.append("versao atual")
    if "noticia" in base.lower() or "diario" in base.lower(): obs.append("fonte secundaria")
    return tipo, muni, "; ".join(obs)

def extrair_texto(path, min_chars=80, ocr=True, zoom=2, log=None):
    doc = fitz.open(path)
    paginas = []; n_ocr = 0
    for i in range(doc.page_count):
        pg = doc[i]
        t = pg.get_text()
        origem = "digital"
        if len(t.strip()) < min_chars and ocr and TEM_OCR:
            try:
                pix = pg.get_pixmap(matrix=fitz.Matrix(zoom, zoom))
                img = Image.open(io.BytesIO(pix.tobytes("png")))
                t = pytesseract.image_to_string(img, lang=os.environ.get("OCR_LANG", "eng"))
                origem = "ocr"; n_ocr += 1
            except Exception as e:
                if log: log(f"   OCR falhou pg {i}: {e}")
        paginas.append((i+1, normalizar(t), origem))
    doc.close()
    return paginas, n_ocr

def detectar_ano(texto_total):
    from collections import Counter
    anos = {}
    for m in re.finditer(r"\b(19|20)\d{2}\b", texto_total):
        a = int(m.group())
        if 1990 <= a <= 2026: anos[a] = anos.get(a,0)+1
    leis_ano = re.findall(r"(?:lei|decreto|lei complementar)[^.]{0,60}?\b(20\d{2}|19\d{2})\b", texto_total)
    info = {"ano_mais_citado": max(anos, key=anos.get) if anos else None,
            "ano_provavel_norma": None,
            "anos_freq": dict(sorted(anos.items(), key=lambda x:-x[1])[:6])}
    if leis_ano:
        info["ano_provavel_norma"] = Counter(int(a) for a in leis_ano).most_common(1)[0][0]
    return info

def detectar_norma(texto_total):
    achados = []
    padroes = {
        "Lei Complementar": r"\blei complementar\b",
        "Lei Ordinaria": r"\blei (?:municipal |n)",
        "Decreto": r"\bdecreto\b",
        "Projeto de Lei": r"\bprojeto de lei\b",
        "Plano (estudo tecnico)": r"\bplano (?:municipal|diretor|de mobilidade|de saneamento)\b",
    }
    for nome, pat in padroes.items():
        if re.search(pat, texto_total): achados.append(nome)
    secs = re.findall(r"secretaria (?:municipal )?(?:de |da |do )?[a-z ,]{3,40}", texto_total)
    secs = [s.strip()[:60] for s in secs][:8]
    empresas = re.findall(r"\b[a-z]{3,}\s+(?:engenharia|consultoria|projetos|servicos|tecnologia|ambiental)\s+(?:ltda|s\.?a\.?|eireli|me)\b", texto_total)
    return {"normas": achados, "secretarias": list(dict.fromkeys(secs)),
            "empresas": list(dict.fromkeys(empresas))[:6]}

def analisar_documento(paginas, P, janela=300):
    ocorr = []
    termos = list(P["termos"]) + [(k, v) for k, v in PRATICA_EXTRA.items()]
    vistos = set()
    for npg, texto, origem in paginas:
        if not texto: continue
        for termo, pat in termos:
            try: rx = re.compile(pat)
            except re.error: continue
            for m in rx.finditer(texto):
                ini = max(0, m.start()-janela); fim = min(len(texto), m.end()+janela)
                ctx = texto[ini:fim]
                h = hashlib.md5((termo+ctx[:160]).encode()).hexdigest()
                if h in vistos: continue
                vistos.add(h)
                terr = sum(1 for w in P["territoriais"] if w in ctx)
                adm = sum(1 for w in P["administrativas"] if w in ctx)
                if termo in PRATICA_EXTRA:
                    forca = 3; tema = "pratica/coleta espacial"; camada = "PRATICA"
                else:
                    forca = P["forca"].get(termo, 1)
                    tema = P["tema"].get(termo, "outros")
                    camada = CAMADA.get(termo, "PERIFERICO")
                ancoragem = 1.0 + min(terr,3)*0.5 - (0.3 if adm > terr else 0)
                ancoragem = max(0.5, ancoragem)
                peso = round(forca*ancoragem, 2)
                ocorr.append({"pagina": npg, "origem": origem, "termo": termo,
                    "tema": tema, "camada": camada, "forca": forca, "terr": terr,
                    "adm": adm, "ancoragem": round(ancoragem,2), "peso": peso,
                    "trecho": m.group(), "contexto": ctx.strip()[:400]})
    return ocorr

def classificar_institucionalizacao(ocorr):
    camadas = set(o["camada"] for o in ocorr if o["camada"] != "PERIFERICO")
    tem_direto = "DIRETO" in camadas
    tem_generico = "GENERICO" in camadas
    tem_pratica = "PRATICA" in camadas
    tem_perif = any(o["camada"] == "PERIFERICO" for o in ocorr)
    if tem_direto: cat = "Explicita"
    elif tem_generico: cat = "Implicita"
    elif tem_pratica: cat = "Pratica (sem terminologia direta)"
    elif tem_perif: cat = "Implicita (indicios perifericos)"
    else: cat = "Ausente"
    return {"categoria": cat, "tem_direto": tem_direto,
            "tem_generico": tem_generico, "tem_pratica": tem_pratica}

def grau_municipio(soma_peso, n_aprov, n_anos, n_temas, tem_forte, rubrica):
    def num(x):
        try: return float(x)
        except Exception: return 0
    for r in rubrica:
        if (r.get("nivel") or "").strip().lower() == "ausente": continue
        ok = (soma_peso >= num(r.get("min_soma_peso")) and
              n_aprov   >= num(r.get("min_ocorr_aproveitaveis")) and
              n_anos    >= num(r.get("min_anos_com_evidencia")) and
              n_temas   >= num(r.get("min_temas_distintos")))
        exige_forte = str(r.get("exige_evidencia_forte","")).strip().lower() in ("sim","s","true")
        if exige_forte and not tem_forte: ok = False
        if ok: return r.get("nivel")
    return "Ausente"

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--pasta", required=True)
    ap.add_argument("--parametros", required=True)
    ap.add_argument("--saida", required=True)
    ap.add_argument("--ocr", default="sim")
    args = ap.parse_args()

    os.makedirs(args.saida, exist_ok=True)
    logpath = os.path.join(args.saida, "progresso.log")
    def log(msg):
        ts = datetime.datetime.now().strftime("%H:%M:%S")
        line = f"[{ts}] {msg}"; print(line, flush=True)
        with open(logpath, "a", encoding="utf-8") as fh: fh.write(line+"\n")

    P = carregar_parametros(args.parametros)
    log(f"Parametros: {len(P['termos'])} padroes, {len(P['forca'])} forcas, "
        f"{len(P['territoriais'])} territoriais. OCR={'on' if TEM_OCR else 'off'} "
        f"FUZZY={'on' if TEM_FUZZY else 'off'}")

    arquivos = sorted(glob.glob(os.path.join(args.pasta, "*.pdf")))
    log(f"{len(arquivos)} PDFs encontrados.")

    resultados = []; todas_ocorr = []
    for k, path in enumerate(arquivos, 1):
        tipo, muni, obs = parse_nome(path)
        log(f"({k}/{len(arquivos)}) {os.path.basename(path)}")
        paginas, n_ocr = extrair_texto(path, ocr=(args.ocr.lower().startswith("s")), log=log)
        texto_total = " ".join(t for _, t, _ in paginas)
        ocorr = analisar_documento(paginas, P)
        info_inst = classificar_institucionalizacao(ocorr)
        ano = detectar_ano(texto_total)
        norma = detectar_norma(texto_total)
        leis = [nome for nome, pat in LEIS_FEDERAIS.items() if re.search(pat, texto_total)]
        soma_peso = round(sum(o["peso"] for o in ocorr), 1)
        temas = sorted(set(o["tema"] for o in ocorr))
        forca_max = max([o["forca"] for o in ocorr], default=0)
        for o in ocorr:
            o2 = dict(o); o2["arquivo"] = os.path.basename(path)
            o2["municipio"] = muni; o2["instrumento"] = tipo
            todas_ocorr.append(o2)
        resultados.append({
            "arquivo": os.path.basename(path), "municipio": muni, "instrumento": tipo,
            "obs": obs, "n_paginas": len(paginas), "paginas_ocr": n_ocr,
            "ano_provavel_norma": ano["ano_provavel_norma"],
            "ano_mais_citado": ano["ano_mais_citado"], "anos_freq": ano["anos_freq"],
            "tipo_norma": "; ".join(norma["normas"]),
            "secretarias": "; ".join(norma["secretarias"][:4]),
            "empresas": "; ".join(norma["empresas"]),
            "categoria_inst": info_inst["categoria"],
            "tem_termo_direto": info_inst["tem_direto"],
            "tem_termo_generico": info_inst["tem_generico"],
            "tem_pratica": info_inst["tem_pratica"],
            "n_ocorrencias": len(ocorr), "soma_peso": soma_peso,
            "forca_max": forca_max, "n_temas": len(temas), "temas": "; ".join(temas),
            "leis_federais": "; ".join(leis) if leis else "(nenhuma)",
            "n_leis_federais": len(leis),
        })
        log(f"   -> {info_inst['categoria']} | ocorr={len(ocorr)} peso={soma_peso} "
            f"leis={len(leis)} ocr_pgs={n_ocr}")

    from collections import defaultdict
    por_muni = defaultdict(list)
    for r in resultados: por_muni[r["municipio"]].append(r)
    graus = {}
    for muni, rs in por_muni.items():
        soma = sum(r["soma_peso"] for r in rs)
        n_aprov = sum(r["n_ocorrencias"] for r in rs)
        anos = set(r["ano_provavel_norma"] for r in rs if r["ano_provavel_norma"])
        temas = set()
        for r in rs:
            if r["temas"]: temas.update(r["temas"].split("; "))
        tem_forte = any(r["forca_max"] >= 3 for r in rs)
        graus[muni] = {"soma_peso": round(soma,1), "n_ocorr": n_aprov,
            "n_anos": len(anos), "n_temas": len([t for t in temas if t]),
            "tem_forte": tem_forte,
            "grau": grau_municipio(soma, n_aprov, len(anos),
                len([t for t in temas if t]), tem_forte, P["rubrica"])}

    with open(os.path.join(args.saida, "resultados.json"), "w", encoding="utf-8") as fh:
        json.dump({"instrumentos": resultados, "graus_municipio": graus,
                   "gerado_em": datetime.datetime.now().isoformat()},
                  fh, ensure_ascii=False, indent=2)
    with open(os.path.join(args.saida, "ocorrencias.json"), "w", encoding="utf-8") as fh:
        json.dump(todas_ocorr, fh, ensure_ascii=False, indent=2)
    log(f"FIM. {len(resultados)} instrumentos, {len(todas_ocorr)} ocorrencias. "
        f"Graus: " + "; ".join(f"{m}={g['grau']}" for m,g in graus.items()))

if __name__ == "__main__":
    main()
