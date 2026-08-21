# -*- coding: utf-8 -*-
# ==================================================================================
# ANALISE LDO/LOA - GEOINFORMACAO MUNICIPAL
# Le os PRINCIPAIS parametros substantivos (vocabulario, forcas, temas, camadas, listas de apoio,
# termos fuzzy, projetos e rubrica do grau) do arquivo Parametros_LDO_LOA.xlsx. Caminhos, regras
# das classes, regex monetaria e heuristicas de OCR permanecem no codigo.
# Roda com OCR (PyMuPDF + Tesseract). Gera saida em Excel com formulas, valores monetarios
# candidatos por trecho e agregacao por projeto/municipio/ano.
# Vocabulario identico aos Anexos A, B e C do Texto-Tese-Modificado.
# ==================================================================================
# COMO USAR (resumo; passo a passo no documento de instrucoes):
#   1) Edite Parametros_LDO_LOA.xlsx (abas de parametros).
#   2) Ajuste, abaixo, ARQ_PARAMETROS e PASTA_RAIZ.
#   3) Rode:  py Codigo_LDO_LOA_v3.py
# Dependencias: pymupdf pandas openpyxl pytesseract pillow rapidfuzz
# ==================================================================================

import os, re, io, sys, unicodedata, subprocess
from pathlib import Path
import pandas as pd

# ================== AREA DE ALTERACAO (somente caminhos) =========================
# Os quatro valores abaixo podem ser sobrescritos por variaveis de ambiente (usado pela
# interface grafica); os defaults abaixo sao os mesmos do script original e continuam
# valendo quando ele e rodado direto (sem a interface), como documentado no README.
ARQ_PARAMETROS = os.environ.get("LDO_ARQ_PARAMETROS", r"C:\analiseldoloa\Parametros_LDO_LOA.xlsx")
PASTA_RAIZ     = os.environ.get("LDO_PASTA_RAIZ", r"C:\analiseldoloa")
ARQ_SAIDA      = os.environ.get("LDO_ARQ_SAIDA", r"C:\analiseldoloa\Resultados_LDO_LOA.xlsx")
CAMINHOS_TESSERACT = [
    r"C:\Program Files\Tesseract-OCR\tesseract.exe",
    r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
]
if os.environ.get("LDO_TESSERACT_PATH"):
    CAMINHOS_TESSERACT.insert(0, os.environ["LDO_TESSERACT_PATH"])
# ================================================================================


def norm(texto):
    if texto is None: return ""
    texto = unicodedata.normalize("NFD", str(texto))
    return "".join(c for c in texto if unicodedata.category(c) != "Mn").lower()

def limpar(texto):
    return re.sub(r"\s+", " ", str(texto or "")).strip()

def slug(texto):
    return re.sub(r"[^a-z0-9]+", "_", norm(texto)).strip("_")


# ==================================================================================
# 1) LEITURA DOS PARAMETROS (Parametros_LDO_LOA.xlsx)
# ==================================================================================
def _sim(v):
    return str(v).strip().lower() in ("sim", "s", "true", "1", "verdadeiro")

def carregar_parametros(caminho):
    import openpyxl
    wb = openpyxl.load_workbook(caminho, data_only=True)

    def linhas(aba):
        ws = wb[aba]
        out = []
        for r in range(3, ws.max_row + 1):  # linha 1 = obs, linha 2 = cabecalho
            vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            if any(v not in (None, "") for v in vals):
                out.append(vals)
        return out

    P = {}
    # Parametros_Gerais
    ger = {str(a).strip(): b for a, b, *_ in linhas("Parametros_Gerais")}
    P["ANO_INICIAL"] = int(ger.get("ANO_INICIAL", 2015))
    P["ANO_FINAL"]   = int(ger.get("ANO_FINAL", 2026))
    P["TIPOS"]       = [t.strip() for t in str(ger.get("TIPOS_DOCUMENTO", "LDO; LOA")).split(";") if t.strip()]
    P["USAR_OCR"]    = _sim(ger.get("USAR_OCR", "Sim"))
    P["IDIOMA_OCR"]  = str(ger.get("IDIOMA_OCR", "por")).strip()
    P["MIN_CHAR"]    = int(ger.get("MIN_CARACTERES_TEXTO_DIGITAL", 80))
    P["FORCAR_OCR"]  = _sim(ger.get("FORCAR_OCR", "Nao"))  # OCR mesmo com texto digital "pobre" (lixo)
    P["OCR_ZOOM"]    = int(ger.get("OCR_ZOOM", 2))
    P["JANELA"]      = int(ger.get("JANELA_CONTEXTO", 300))
    P["USAR_FUZZY"]  = _sim(ger.get("USAR_FUZZY", "Sim"))
    P["FUZZY_LIMIAR"]= int(ger.get("FUZZY_LIMIAR", 90))
    P["FUZZY_SO_OCR"]= _sim(ger.get("FUZZY_SO_EM_PAGINAS_OCR", "Sim"))
    P["DEDUP"]       = _sim(ger.get("DEDUP_BOILERPLATE", "Sim"))
    P["DEDUP_HASH"]  = int(ger.get("DEDUP_TAM_HASH", 160))
    P["EXTRAIR_ORC"] = _sim(ger.get("EXTRAIR_ORCAMENTO", "Sim"))

    # Termos_Busca: termo_chave -> [regex...]
    tb = {}
    for chave, padrao, *_ in linhas("Termos_Busca"):
        if chave is None or padrao is None: continue
        tb.setdefault(str(chave).strip(), []).append(str(padrao).strip())
    P["TERMOS_BUSCA"] = tb

    # Tema_Padrao
    P["TEMA"] = {str(a).strip(): str(b).strip() for a, b, *_ in linhas("Tema_Padrao") if a}
    # Forca_Tecnica
    P["FORCA"] = {str(a).strip(): int(b) for a, b, *_ in linhas("Forca_Tecnica") if a and b not in (None, "")}
    try:
        P["CAMADA"] = {str(a).strip(): str(b).strip() for a, b, *_ in linhas("Camada") if a}
    except Exception:
        P["CAMADA"] = {}

    # Vocabulario_Apoio: lista -> [valores]
    voc = {}
    for lista, valor, *_ in linhas("Vocabulario_Apoio"):
        if lista and valor: voc.setdefault(str(lista).strip(), []).append(norm(valor))
    P["TERRITORIAIS"] = voc.get("PALAVRAS_TERRITORIAIS", [])
    P["ADMIN"]        = voc.get("PALAVRAS_ADMINISTRATIVAS", [])
    P["FALSO_POS"]    = voc.get("EXPRESSOES_FALSO_POSITIVO", [])
    P["FISCAL_FRACO"] = voc.get("EXPRESSOES_CADASTRO_FISCAL_FRACO", [])

    # Termos_Canon_Fuzzy
    P["FUZZY_TERMOS"] = [norm(a) for a, *_ in linhas("Termos_Canon_Fuzzy") if a]

    # Grau_Rubrica
    grau = {}
    for nivel, sp, oa, ae, td, ef, *_ in linhas("Grau_Rubrica"):
        if not nivel or str(nivel).strip() == "Ausente": continue
        def _i(x):
            try: return int(x)
            except Exception: return 0
        grau[str(nivel).strip()] = {"soma": _i(sp), "aprov": _i(oa), "anos": _i(ae),
                                    "temas": _i(td), "forte": _sim(ef)}
    P["GRAU"] = grau

    # Projetos: termo_chave -> projeto ; e metadados do projeto
    proj_termo = {}
    proj_meta = {}
    for row in linhas("Projetos"):
        row = list(row) + ["", "", "", "", ""]
        projeto, eixo, aderencia, termos, interpret = row[0], row[1], row[2], row[3], row[4]
        if not projeto: continue
        # coluna de termos: usa a que tiver ";" (robusto a formatos de 2 ou 5 colunas)
        if not (termos and ";" in str(termos)):
            for cand in row[1:]:
                if cand and ";" in str(cand): termos = cand; break
            else:
                termos = row[1] if len(row) > 1 else ""
        proj_meta[str(projeto).strip()] = {"eixo": eixo, "aderencia": aderencia, "interpret": interpret}
        for t in str(termos or "").split(";"):
            t = t.strip()
            if t: proj_termo[norm(t)] = str(projeto).strip()
    P["PROJ_TERMO"] = proj_termo
    P["PROJ_META"] = proj_meta
    return P


# ==================================================================================
# 2) OCR / EXTRACAO DE TEXTO (PyMuPDF -> pdftotext) + Tesseract
# ==================================================================================
try:
    import fitz
    TEM_FITZ = True
except Exception:
    TEM_FITZ = False

OCR_OK = False
def _config_ocr(P):
    global OCR_OK
    if not P["USAR_OCR"]: return
    try:
        import pytesseract
        from PIL import Image  # noqa
        for c in CAMINHOS_TESSERACT:
            if os.path.exists(c):
                pytesseract.pytesseract.tesseract_cmd = c; OCR_OK = True; break
        if not OCR_OK:
            subprocess.run(["tesseract", "--version"], capture_output=True); OCR_OK = True
    except Exception:
        OCR_OK = False
        print("AVISO: OCR indisponivel (instale pytesseract/pillow e o Tesseract).")

def _pdftotext(caminho):
    try:
        out = subprocess.run(["pdftotext", "-layout", str(caminho), "-"], capture_output=True, timeout=600)
        txt = out.stdout.decode("utf-8", "ignore")
    except Exception:
        return []
    return [{"pagina": i+1, "texto": p, "origem_texto": "texto digital"} for i, p in enumerate(txt.split("\f"))]

PALAVRAS_COMUNS = [" de "," da "," do "," das "," dos "," para "," com "," secretaria ",
                   " prefeitura "," municipio "," lei "," valor "," total "," programa ",
                   " orcamento "," despesa "," receita "," art "," paragrafo "]

def texto_pobre(td):
    """True quando a pagina TEM texto, mas o texto parece lixo (camada de OCR ruim),
    isto e: poucas palavras comuns do portugues/orcamento por palavra."""
    tn = " " + norm(td) + " "
    w = tn.split()
    if len(w) < 30:
        return False  # paginas curtas (capa/cabecalho) nao sao reprocessadas
    comuns = sum(tn.count(x) for x in PALAVRAS_COMUNS)
    return (comuns / len(w)) < 0.05

def extrair_texto(caminho, P):
    if TEM_FITZ:
        res = []
        doc = fitz.open(str(caminho))
        for i in range(len(doc)):
            pag = doc[i]; td = pag.get_text("text") or ""; origem, final = "texto digital", td
            precisa_ocr = len(limpar(td)) < P["MIN_CHAR"] or (P["FORCAR_OCR"] and texto_pobre(td))
            if precisa_ocr and P["USAR_OCR"] and OCR_OK:
                try:
                    import pytesseract
                    from PIL import Image
                    pix = pag.get_pixmap(matrix=fitz.Matrix(P["OCR_ZOOM"], P["OCR_ZOOM"]), alpha=False)
                    img = Image.open(io.BytesIO(pix.tobytes("png")))
                    tocr = pytesseract.image_to_string(img, lang=P["IDIOMA_OCR"])
                    if len(limpar(tocr)) > len(limpar(td)): origem, final = "OCR", tocr
                except Exception as e:
                    origem = "erro OCR: %s" % e
            res.append({"pagina": i+1, "texto": final, "origem_texto": origem})
        doc.close(); return res
    return _pdftotext(caminho)


# ==================================================================================
# 3) FUZZY (recupera termos corrompidos por OCR)
# ==================================================================================
try:
    from rapidfuzz import fuzz as _fz
except Exception:
    import difflib
    class _FZ:
        @staticmethod
        def ratio(a, b): return difflib.SequenceMatcher(None, a, b).ratio() * 100
    _fz = _FZ()


# ==================================================================================
# 4) CLASSIFICACAO EM DOIS EIXOS + DOTACAO + PROJETO
# ==================================================================================
RE_VAL = re.compile(r"\d{1,3}(?:\.\d{3})+,\d{2}")
RE_VAL2 = re.compile(r"\b\d+,\d{2}\b")
RE_COD = re.compile(r"\b\d{2}\.\d{3}\b|\b\d\.\d{3}\b")

def valor_para_numero(s):
    s = str(s).strip()
    if not s: return None
    try: return float(s.replace(".", "").replace(",", "."))
    except Exception: return None

def ancoragem(termo_norm, trecho_norm, P, encontrado_norm=""):
    if any(e in trecho_norm for e in P["FALSO_POS"]): return 0
    # remove o PROPRIO termo do trecho para nao se autoconfirmar como territorial
    trecho_terr = trecho_norm
    for rem in {termo_norm, encontrado_norm}:
        if rem: trecho_terr = trecho_terr.replace(rem, " ")
    tem_terr = any(p in trecho_terr for p in P["TERRITORIAIS"])
    tem_adm  = any(p in trecho_norm for p in P["ADMIN"])
    nucleo = P["FORCA"].get(_chave(termo_norm, P), 0) >= 3
    if termo_norm == "cadastro social":
        return 1 if tem_terr else 0
    if termo_norm == "cadastro fiscal":
        if any(e in trecho_norm for e in P["FISCAL_FRACO"]): return 1
        return 2 if tem_terr else 1
    if termo_norm == "mapeamento":
        if tem_adm and not tem_terr: return 0
        return 2 if tem_terr else 1
    if tem_terr: return 2
    if tem_adm and not nucleo: return 0
    return 2 if nucleo else 1

def _chave(termo_norm, P):
    if not hasattr(_chave, "_m") or _chave._src is not P:
        _chave._m = {norm(k): k for k in P["FORCA"]}; _chave._src = P
    return _chave._m.get(termo_norm, termo_norm)

def classe_textual(forca, anc):
    if anc == 0: return "ocorrencia administrativa generica / falso positivo"
    if anc >= 2 and forca >= 2: return "ocorrencia territorial relevante"
    if forca >= 3 and anc >= 1: return "ocorrencia validada (tecnica)"
    return "ocorrencia contextual"

def descricao_acao(trecho, termo_encontrado):
    """Captura o rotulo de acao apos o termo, ex.: 'NUCLEO DE GEOPROCESSAMENTO - <projeto>'."""
    m = re.search(re.escape(termo_encontrado) + r"\s*[-:]\s*([A-Za-zÀ-ÿ /]{6,80})", trecho)
    if m: return limpar(m.group(1))[:80]
    m = re.search(r"(?:ACAO|AÇÃO|PROJETO|ATIVIDADE)\s*[:\-]\s*([A-Za-zÀ-ÿ /]{6,80})", trecho, re.IGNORECASE)
    return limpar(m.group(1))[:80] if m else ""

def dotacao_proxima(trecho):
    m = RE_VAL.findall(trecho)
    if m: return m[0]
    m = RE_VAL2.findall(trecho)
    return m[0] if m else ""


# ==================================================================================
# 5) BUSCA DE TERMOS (exato + fuzzy, dedup por sobreposicao)
# ==================================================================================
def _ov(a0, a1, b0, b1): return max(a0, b0) < min(a1, b1)

def buscar(texto, origem, P):
    tnorm = norm(texto or "")
    cand = []
    for chave, padroes in P["TERMOS_BUSCA"].items():
        for pad in padroes:
            try:
                matches = list(re.finditer(norm(pad), tnorm))
            except re.error as e:
                print("AVISO: padrao regex invalido ignorado (%s): %s -> %s" % (chave, pad, e))
                continue
            for mm in matches:
                cand.append([chave, mm.group(0), mm.start(), mm.end(), mm.end()-mm.start(), "exato"])
    if P["USAR_FUZZY"] and (origem == "OCR" or not P["FUZZY_SO_OCR"]):
        for alvo in P["FUZZY_TERMOS"]:
            if " " in alvo: continue
            for tk in re.finditer(r"[a-z]{%d,}" % max(7, len(alvo)-3), tnorm):
                if abs(len(tk.group(0)) - len(alvo)) > 4: continue
                if _fz.ratio(tk.group(0), alvo) >= P["FUZZY_LIMIAR"]:
                    cand.append([_chave(alvo, P), tk.group(0), tk.start(), tk.end(), tk.end()-tk.start(), "fuzzy"])
    cand.sort(key=lambda x: (x[2], -x[4]))
    sel = []
    for c in cand:
        conf = False
        for s in list(sel):
            if _ov(c[2], c[3], s[2], s[3]):
                conf = True
                if c[4] > s[4]: sel.remove(s); sel.append(c)
                break
        if not conf: sel.append(c)
    sel.sort(key=lambda x: x[2])
    import hashlib
    out = []
    for chave, enc, ini, fim, _t, metodo in sel:
        a = max(0, ini - P["JANELA"]); b = min(len(texto), fim + P["JANELA"])
        trecho = limpar(texto[a:b]); tnz = norm(trecho)
        forca = P["FORCA"].get(chave, 1); anc = ancoragem(norm(chave), tnz, P, norm(enc))
        peso = forca * anc
        projeto = P["PROJ_TERMO"].get(norm(chave), "")
        out.append({
            "termo_chave": chave, "termo_encontrado": enc, "metodo_busca": metodo,
            "trecho": trecho, "hash_trecho": hashlib.md5(tnz[:P["DEDUP_HASH"]].encode()).hexdigest(),
            "forca_tecnica": forca, "ancoragem_territorial": anc, "peso_analitico": peso,
            "tema_analitico": P["TEMA"].get(chave, "outros"), "camada": P.get("CAMADA",{}).get(chave, ""), "classe": classe_textual(forca, anc),
            "projeto": projeto, "descricao_acao": descricao_acao(trecho, enc),
            "valor_dotacao": dotacao_proxima(trecho) if P["EXTRAIR_ORC"] else "",
        })
    return out


# ==================================================================================
# 6) LOCALIZAR ARQUIVOS / PROCESSAR
# ==================================================================================
def nome_municipio(nome):
    n = nome.strip()
    for p in [r"\s+LOA\s+E\s+LDO$", r"\s+LDO\s+E\s+LOA$"]:
        n = re.sub(p, "", n, flags=re.IGNORECASE)
    return n.strip()

def listar_pastas(raiz):
    raiz = Path(raiz)
    if not raiz.exists(): raise FileNotFoundError("Pasta raiz nao encontrada: %s" % raiz)
    return [it for it in sorted(raiz.iterdir()) if it.is_dir()
            and "loa" in norm(it.name) and "ldo" in norm(it.name)
            and not norm(it.name).startswith("saida")]

def localizar_pdf(pasta, ano, tipo):
    c = [a for a in Path(pasta).rglob("*.pdf") if str(ano) in norm(a.name) and norm(tipo) in norm(a.name)]
    return sorted(c)[0] if c else None

def processar_tudo(P):
    anos = list(range(P["ANO_FINAL"], P["ANO_INICIAL"]-1, -1))
    resumo, ocorr = [], []
    for pasta in listar_pastas(PASTA_RAIZ):
        municipio = nome_municipio(pasta.name)
        print("\n===", municipio, "===")
        for ano in anos:
            for tipo in P["TIPOS"]:
                pdf = localizar_pdf(pasta, ano, tipo)
                if pdf is None:
                    resumo.append({"municipio": municipio, "ano": ano, "tipo_documento": tipo,
                                   "arquivo": "", "status": "documento ausente", "total_paginas": 0,
                                   "paginas_texto_digital": 0, "paginas_ocr": 0, "paginas_erro_ocr": 0})
                    continue
                try:
                    paginas = extrair_texto(pdf, P)
                except Exception as e:
                    resumo.append({"municipio": municipio, "ano": ano, "tipo_documento": tipo,
                                   "arquivo": pdf.name, "status": "erro: %s" % e, "total_paginas": 0,
                                   "paginas_texto_digital": 0, "paginas_ocr": 0, "paginas_erro_ocr": 0})
                    continue
                nd = sum(1 for p in paginas if p["origem_texto"] == "texto digital")
                no = sum(1 for p in paginas if p["origem_texto"] == "OCR")
                ne = sum(1 for p in paginas if str(p["origem_texto"]).startswith("erro"))
                for p in paginas:
                    for o in buscar(p["texto"], p["origem_texto"], P):
                        o.update({"municipio": municipio, "ano": ano, "tipo_documento": tipo,
                                  "arquivo": pdf.name, "pagina": p["pagina"], "origem_texto": p["origem_texto"]})
                        ocorr.append(o)
                resumo.append({"municipio": municipio, "ano": ano, "tipo_documento": tipo,
                               "arquivo": pdf.name, "status": "processado", "total_paginas": len(paginas),
                               "paginas_texto_digital": nd, "paginas_ocr": no, "paginas_erro_ocr": ne})
                print("  %s %s: %d ocorrencias (%d paginas, %d OCR)" %
                      (tipo, ano, sum(1 for x in ocorr if x.get("ano")==ano and x.get("tipo_documento")==tipo and x.get("municipio")==municipio), len(paginas), no))
    df_oc = pd.DataFrame(ocorr)
    if P["DEDUP"] and not df_oc.empty:
        df_oc = df_oc.drop_duplicates(subset=["municipio", "arquivo", "hash_trecho"]).reset_index(drop=True)
    return pd.DataFrame(resumo), df_oc


# ==================================================================================
# 7) GRAU DE INTEGRACAO (rubrica vinda dos parametros)
# ==================================================================================
def classificar_grau(g, P):
    if g is None or g.empty: return "Ausente", {}
    d = g[g["peso_analitico"] > 0]
    if d.empty: return "Ausente", {}
    met = {"soma_peso": int(d["peso_analitico"].sum()),
           "ocorr_aproveitaveis": int(len(d)),
           "anos_com_evidencia": int(d["ano"].nunique()),
           "temas_distintos": int(d["tema_analitico"].nunique()),
           "tem_evidencia_forte": bool((d["forca_tecnica"] >= 3).any())}
    for nivel in ["Consolidado", "Parcial", "Incipiente"]:
        r = P["GRAU"].get(nivel)
        if not r: continue
        if (met["soma_peso"] >= r["soma"] and met["ocorr_aproveitaveis"] >= r["aprov"]
                and met["anos_com_evidencia"] >= r["anos"] and met["temas_distintos"] >= r["temas"]
                and (met["tem_evidencia_forte"] or not r["forte"])):
            return nivel, met
    return "Ausente", met


# ==================================================================================
# 8) SAIDA NO FORMATO DASHBOARD, COM FORMULAS
# ==================================================================================
def build_saida(df_resumo, df_oc, P, caminho):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference
    from openpyxl.utils import get_column_letter

    hdr_fill = PatternFill("solid", fgColor="D5E8F0")
    hdr_font = Font(bold=True, color="1F3864")
    title_font = Font(bold=True, size=16, color="1F3864")
    thin = Side(style="thin", color="CCCCCC")
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")

    df_oc = df_oc.copy()
    if df_oc.empty:
        df_oc = pd.DataFrame(columns=["municipio","ano","tipo_documento","pagina","termo_chave",
            "termo_encontrado","tema_analitico","forca_tecnica","ancoragem_territorial","peso_analitico",
            "classe","valor_dotacao","projeto","descricao_acao","origem_texto","trecho"])
    df_oc["valor_num"] = df_oc["valor_dotacao"].map(valor_para_numero) if "valor_dotacao" in df_oc else None

    anos = sorted(df_resumo["ano"].unique())
    municipios = sorted(df_resumo["municipio"].unique())
    n_tipos = len(P["TIPOS"]); n_anos = len(anos)
    esperados_por_mun = n_tipos * n_anos
    esperados_por_ano = n_tipos * len(municipios)

    wb = openpyxl.Workbook()

    def estilizar(ws, ncol, header_row=1, freeze=True):
        for j in range(1, ncol+1):
            c = ws.cell(header_row, j); c.font = hdr_font; c.fill = hdr_fill; c.border = bd
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if freeze: ws.freeze_panes = ws.cell(header_row+1, 1)

    def escrever_df(ws, df, start=1):
        for j, col in enumerate(df.columns, 1):
            ws.cell(start, j, col)
        for i, (_, row) in enumerate(df.iterrows(), start+1):
            for j, col in enumerate(df.columns, 1):
                v = row[col]
                ws.cell(i, j, None if pd.isna(v) else v)
        estilizar(ws, len(df.columns), start)

    # ---------- Ocorrencias_Relevantes (dados + formulas peso/aproveitavel) ----------
    ws = wb.active; ws.title = "Ocorrencias_Relevantes"
    cols = ["municipio","ano","tipo_documento","pagina","termo_chave","termo_encontrado",
            "tema_analitico","forca_tecnica","ancoragem_territorial","peso_analitico","aproveitavel",
            "classe","valor_dotacao_R$","projeto","descricao_acao","origem_texto","trecho","camada"]
    for j,c in enumerate(cols,1): ws.cell(1,j,c)
    base = df_oc.sort_values(["municipio","ano","tipo_documento"]) if not df_oc.empty else df_oc
    r = 2
    for _, o in base.iterrows():
        ws.cell(r,1,o["municipio"]); ws.cell(r,2,int(o["ano"])); ws.cell(r,3,o["tipo_documento"])
        ws.cell(r,4,int(o.get("pagina",0)) if not pd.isna(o.get("pagina")) else 0)
        ws.cell(r,5,o["termo_chave"]); ws.cell(r,6,o.get("termo_encontrado",""))
        ws.cell(r,7,o.get("tema_analitico","")); ws.cell(r,8,int(o["forca_tecnica"]))
        ws.cell(r,9,int(o["ancoragem_territorial"]))
        ws.cell(r,10,"=H%d*I%d" % (r, r))            # peso = forca x ancoragem (FORMULA)
        ws.cell(r,11,"=IF(J%d>0,1,0)" % r)            # aproveitavel (FORMULA)
        ws.cell(r,12,o.get("classe",""))
        ws.cell(r,13, o.get("valor_num") if not pd.isna(o.get("valor_num")) else None)
        ws.cell(r,14,o.get("projeto","")); ws.cell(r,15,o.get("descricao_acao",""))
        ws.cell(r,16,o.get("origem_texto","")); ws.cell(r,17,str(o.get("trecho",""))[:600]); ws.cell(r,18,o.get("camada",""))
        r += 1
    LR = max(r-1, 2)
    estilizar(ws, len(cols))
    ws.column_dimensions["Q"].width = 70
    R = "'Ocorrencias_Relevantes'!"
    G = lambda col: "%s$%s$2:$%s$%d" % (R, col, col, LR)   # range coluna

    # ---------- Resumo_Documentos ----------
    ws = wb.create_sheet("Resumo_Documentos")
    rd = df_resumo[["municipio","ano","tipo_documento","arquivo","status","total_paginas",
                    "paginas_texto_digital","paginas_ocr","paginas_erro_ocr"]].sort_values(["municipio","ano","tipo_documento"])
    escrever_df(ws, rd)
    RDlast = len(rd) + 1
    RD = "'Resumo_Documentos'!"
    RDc = lambda col: "%s$%s$2:$%s$%d" % (RD, col, col, RDlast)

    # ---------- Resumo_Municipios (FORMULAS) ----------
    ws = wb.create_sheet("Resumo_Municipios")
    head = ["Município","Documentos esperados","Documentos processados","Documentos ausentes",
            "Cobertura documental","Ocorrências brutas","Ocorrências aproveitáveis","Soma do peso analítico",
            "Peso por doc processado","Participação nas ocorrências"]
    for j,h in enumerate(head,1): ws.cell(1,j,h)
    total_brutas_ref = "SUMPRODUCT(--(%s<>\"\"))" % G("A")
    for i, mun in enumerate(municipios, 2):
        ws.cell(i,1,mun)
        ws.cell(i,2,esperados_por_mun)
        ws.cell(i,3,'=COUNTIFS(%s,A%d,%s,"processado")' % (RDc("A"), i, RDc("E")))
        ws.cell(i,4,"=B%d-C%d" % (i,i))
        ws.cell(i,5,"=IFERROR(C%d/B%d,0)" % (i,i))
        ws.cell(i,6,"=COUNTIFS(%s,A%d)" % (G("A"), i))
        ws.cell(i,7,"=SUMIFS(%s,%s,A%d)" % (G("K"), G("A"), i))
        ws.cell(i,8,"=SUMIFS(%s,%s,A%d)" % (G("J"), G("A"), i))
        ws.cell(i,9,"=IFERROR(H%d/C%d,0)" % (i,i))
        ws.cell(i,10,"=IFERROR(F%d/%s,0)" % (i, total_brutas_ref))
    estilizar(ws, len(head))
    RM_last = len(municipios)+1

    # ---------- Evolucao_Anual (FORMULAS) ----------
    ws = wb.create_sheet("Evolucao_Anual")
    for j,h in enumerate(["Ano","Ocorrências brutas","Ocorrências aproveitáveis","Soma do peso analítico","Documentos processados","Documentos ausentes"],1):
        ws.cell(1,j,h)
    for i, ano in enumerate(anos, 2):
        ws.cell(i,1,ano)
        ws.cell(i,2,"=COUNTIFS(%s,A%d)" % (G("B"), i))
        ws.cell(i,3,"=SUMIFS(%s,%s,A%d)" % (G("K"), G("B"), i))
        ws.cell(i,4,"=SUMIFS(%s,%s,A%d)" % (G("J"), G("B"), i))
        ws.cell(i,5,'=COUNTIFS(%s,A%d,%s,"processado")' % (RDc("B"), i, RDc("E")))
        ws.cell(i,6,"=%d-E%d" % (esperados_por_ano, i))
    estilizar(ws, 6)

    # ---------- Evolucao_Municipio_Ano (valores) ----------
    rows=[]
    for mun in municipios:
        for ano in anos:
            sub = df_oc[(df_oc["municipio"]==mun)&(df_oc["ano"]==ano)] if not df_oc.empty else df_oc
            rsub = df_resumo[(df_resumo["municipio"]==mun)&(df_resumo["ano"]==ano)]
            def cnt(t): return int(len(sub[sub["tipo_documento"]==t])) if not sub.empty else 0
            def pes(t): return int(sub[sub["tipo_documento"]==t]["peso_analitico"].sum()) if not sub.empty else 0
            st = {t: (rsub[rsub["tipo_documento"]==t]["status"].iloc[0] if len(rsub[rsub["tipo_documento"]==t]) else "ausente") for t in P["TIPOS"]}
            rows.append({"Município":mun,"Ano":ano,"Ocorrências brutas":int(len(sub)) if not sub.empty else 0,
                "Ocorrências aproveitáveis":int((sub["peso_analitico"]>0).sum()) if not sub.empty else 0,
                "Soma do peso analítico":int(sub["peso_analitico"].sum()) if not sub.empty else 0,
                "Ocorrências LDO":cnt("LDO"),"Ocorrências LOA":cnt("LOA"),"Peso LDO":pes("LDO"),"Peso LOA":pes("LOA"),
                "Status LDO":st.get("LDO",""),"Status LOA":st.get("LOA","")})
    ws = wb.create_sheet("Evolucao_Municipio_Ano"); escrever_df(ws, pd.DataFrame(rows))

    # ---------- Temas (FORMULAS) ----------
    temas = sorted(df_oc["tema_analitico"].dropna().unique()) if not df_oc.empty else []
    ws = wb.create_sheet("Temas")
    for j,h in enumerate(["Tema analítico","Ocorrências brutas","Ocorrências aproveitáveis","Soma do peso analítico","Municípios"],1): ws.cell(1,j,h)
    for i, tm in enumerate(temas, 2):
        ws.cell(i,1,tm)
        ws.cell(i,2,"=COUNTIFS(%s,A%d)" % (G("G"), i))
        ws.cell(i,3,"=SUMIFS(%s,%s,A%d)" % (G("K"), G("G"), i))
        ws.cell(i,4,"=SUMIFS(%s,%s,A%d)" % (G("J"), G("G"), i))
        muns = ", ".join(sorted(df_oc[df_oc["tema_analitico"]==tm]["municipio"].unique()))
        ws.cell(i,5,muns)
    estilizar(ws, 5)

    # ---------- Termos (FORMULAS) ----------
    termos = sorted(df_oc["termo_chave"].dropna().unique()) if not df_oc.empty else []
    ws = wb.create_sheet("Termos")
    for j,h in enumerate(["Termo-chave","Ocorrências brutas","Ocorrências aproveitáveis","Soma do peso analítico","Municípios"],1): ws.cell(1,j,h)
    for i, tm in enumerate(termos, 2):
        ws.cell(i,1,tm)
        ws.cell(i,2,"=COUNTIFS(%s,A%d)" % (G("E"), i))
        ws.cell(i,3,"=SUMIFS(%s,%s,A%d)" % (G("K"), G("E"), i))
        ws.cell(i,4,"=SUMIFS(%s,%s,A%d)" % (G("J"), G("E"), i))
        muns = ", ".join(sorted(df_oc[df_oc["termo_chave"]==tm]["municipio"].unique()))
        ws.cell(i,5,muns)
    estilizar(ws, 5)

    # ---------- Melhores_Projetos (FORMULAS p/ ocorr, peso e DOTACAO) ----------
    projs = sorted([p for p in df_oc["projeto"].dropna().unique() if str(p).strip()]) if not df_oc.empty else []
    ws = wb.create_sheet("Melhores_Projetos")
    for j,h in enumerate(["Projeto/ação","Eixo","Aderência","Ocorrências","Soma do peso","Dotação total (R$)","Municípios","Anos","Interpretação"],1): ws.cell(1,j,h)
    for i, pj in enumerate(projs, 2):
        meta = P["PROJ_META"].get(pj, {})
        ws.cell(i,1,pj); ws.cell(i,2,meta.get("eixo","")); ws.cell(i,3,meta.get("aderencia",""))
        ws.cell(i,4,"=COUNTIFS(%s,A%d)" % (G("N"), i))
        ws.cell(i,5,"=SUMIFS(%s,%s,A%d)" % (G("J"), G("N"), i))
        ws.cell(i,6,"=SUMIFS(%s,%s,A%d)" % (G("M"), G("N"), i))
        sub = df_oc[df_oc["projeto"]==pj]
        ws.cell(i,7,", ".join(sorted(sub["municipio"].unique())))
        ws.cell(i,8,", ".join(map(str, sorted(sub["ano"].unique()))))
        ws.cell(i,9,meta.get("interpret",""))
    estilizar(ws, 9)

    # ---------- Projetos_Municipio_Ano (valores) - "como Duque para cada municipio" ----------
    rows=[]
    if not df_oc.empty:
        gp = df_oc[df_oc["projeto"].astype(str).str.strip()!=""]
        for (mun,pj,ano), s in gp.groupby(["municipio","projeto","ano"]):
            rows.append({"Município":mun,"Projeto/ação":pj,"Ano":int(ano),"Menções":int(len(s)),
                         "Soma do peso":int(s["peso_analitico"].sum()),
                         "Dotação estimada (R$)":float(s["valor_num"].dropna().sum()) if s["valor_num"].notna().any() else None,
                         "Descrições capturadas":"; ".join(sorted(set([d for d in s["descricao_acao"] if str(d).strip()]))[:3])})
    ws = wb.create_sheet("Projetos_Municipio_Ano")
    escrever_df(ws, pd.DataFrame(rows) if rows else pd.DataFrame(columns=["Município","Projeto/ação","Ano","Menções","Soma do peso","Dotação estimada (R$)","Descrições capturadas"]))

    # ---------- Grau_Integracao (valores) ----------
    rows=[]
    for mun in municipios:
        g = df_oc[df_oc["municipio"]==mun] if not df_oc.empty else df_oc
        nivel, met = classificar_grau(g, P)
        rows.append({"Município":mun,"Grau de Integração":nivel,
                     "Soma do peso":met.get("soma_peso",0),"Aproveitáveis":met.get("ocorr_aproveitaveis",0),
                     "Anos com evidência":met.get("anos_com_evidencia",0),"Temas distintos":met.get("temas_distintos",0),
                     "Tem evidência forte":("Sim" if met.get("tem_evidencia_forte") else "Não")})
    ws = wb.create_sheet("Grau_Integracao"); escrever_df(ws, pd.DataFrame(rows))

    # ---------- Dashboard (KPIs em formula + graficos) ----------
    ws = wb.create_sheet("Dashboard", 0)
    ws["A1"] = "Relatório - LDO/LOA Baixada Fluminense Histórica (v3)"; ws["A1"].font = title_font
    ws["A2"] = "Período: %d-%d | Parâmetros: Parametros_LDO_LOA.xlsx | Fórmulas recalculam ao editar." % (P["ANO_INICIAL"], P["ANO_FINAL"])
    rmN = RM_last
    kpis = [
        ("Municípios analisados", "=COUNTA('Resumo_Municipios'!$A$2:$A$%d)" % rmN),
        ("Documentos esperados", "=SUM('Resumo_Municipios'!$B$2:$B$%d)" % rmN),
        ("Documentos processados", "=SUM('Resumo_Municipios'!$C$2:$C$%d)" % rmN),
        ("Cobertura documental", "=IFERROR(B7/B6,0)"),
        ("Ocorrências brutas", "=SUM('Resumo_Municipios'!$F$2:$F$%d)" % rmN),
        ("Ocorrências aproveitáveis", "=SUM('Resumo_Municipios'!$G$2:$G$%d)" % rmN),
        ("Soma do peso analítico", "=SUM('Resumo_Municipios'!$H$2:$H$%d)" % rmN),
    ]
    ws["A4"]="Indicador"; ws["B4"]="Resultado"; ws["A4"].font=hdr_font; ws["B4"].font=hdr_font
    ws["A4"].fill=hdr_fill; ws["B4"].fill=hdr_fill
    for k,(rotulo,formula) in enumerate(kpis, 5):
        ws.cell(k,1,rotulo); ws.cell(k,2,formula)
    # graficos
    ch1 = BarChart(); ch1.title="Soma do peso por município"; ch1.type="bar"
    data = Reference(wb["Resumo_Municipios"], min_col=8, min_row=1, max_row=rmN)
    cats = Reference(wb["Resumo_Municipios"], min_col=1, min_row=2, max_row=rmN)
    ch1.add_data(data, titles_from_data=True); ch1.set_categories(cats); ch1.height=8; ch1.width=16
    ws.add_chart(ch1, "D4")
    ch2 = BarChart(); ch2.title="Soma do peso por ano"
    ea_last = len(anos)+1
    data2 = Reference(wb["Evolucao_Anual"], min_col=4, min_row=1, max_row=ea_last)
    cats2 = Reference(wb["Evolucao_Anual"], min_col=1, min_row=2, max_row=ea_last)
    ch2.add_data(data2, titles_from_data=True); ch2.set_categories(cats2); ch2.height=8; ch2.width=16
    ws.add_chart(ch2, "D24")
    ws.column_dimensions["A"].width=30; ws.column_dimensions["B"].width=16

    # ---------- Metodologia ----------
    ws = wb.create_sheet("Metodologia")
    met = [("Base","Consolidado das LDOs e LOAs, %d-%d, dos municípios da Baixada Fluminense Histórica." % (P["ANO_INICIAL"],P["ANO_FINAL"])),
           ("Parâmetros","Definidos em Parametros_LDO_LOA.xlsx (termos, pesos, rubrica, projetos)."),
           ("Peso analítico","peso = força técnica (1-3) x ancoragem territorial (0-2). Célula com fórmula."),
           ("Ocorrência aproveitável","Ocorrência com peso > 0."),
           ("Dotação","Valor em R$ mais próximo do termo no trecho (conferir; sujeito a OCR)."),
           ("Projeto","Mapa termo->projeto da aba Projetos; dotação somada por projeto."),
           ("Grau de Integração","Rubrica da aba Grau_Rubrica, por município, sobre toda a série."),
           ("Fonte","Elaboração própria a partir das LDOs/LOAs municipais.")]
    escrever_df(ws, pd.DataFrame(met, columns=["Item","Descrição"]))

    wb.save(caminho)
    print("Saida salva em:", caminho)


def main():
    print("Lendo parametros:", ARQ_PARAMETROS)
    P = carregar_parametros(ARQ_PARAMETROS)
    _config_ocr(P)
    print("Backend:", "PyMuPDF" if TEM_FITZ else "pdftotext", "| OCR:", "ON" if (P["USAR_OCR"] and OCR_OK) else "OFF")
    df_resumo, df_oc = processar_tudo(P)
    build_saida(df_resumo, df_oc, P, ARQ_SAIDA)
    print("\nConcluido. Ocorrencias:", 0 if df_oc is None else len(df_oc))


if __name__ == "__main__":
    main()
