# -*- coding: utf-8 -*-
# ==================================================================================
# FERRAMENTAS AUXILIARES - LDO/LOA
# ==================================================================================
# Dois utilitarios:
#
# 1) PREENCHER por correspondencia: copia os rotulos ja codificados de uma amostra
#    para outra amostra que tenha as MESMAS ocorrencias (mesmo municipio, ano, tipo,
#    pagina e termo). Util para aproveitar a amostra_para_codificar_200 ja preenchida
#    na amostra_validacao_v3.
#
#    py Ferramentas_LDO_LOA.py preencher  <amostra_PREENCHIDA.xlsx>  <amostra_ALVO.xlsx>  <saida.xlsx>
#
# 2) REMOVER termos do arquivo de parametros (ex.: cadastro social).
#
#    py Ferramentas_LDO_LOA.py remover  Parametros_LDO_LOA.xlsx  "cadastro social"  ["outro termo" ...]
#
# Dependencias: pandas, openpyxl
# ==================================================================================
import sys, unicodedata
import pandas as pd
import openpyxl

def norm(s):
    s = unicodedata.normalize("NFD", str(s))
    return "".join(c for c in s if unicodedata.category(c) != "Mn").lower().strip()

def chave(row):
    # chave robusta: municipio | ano | tipo | pagina | termo_chave | termo_encontrado
    return "|".join([norm(row.get("municipio","")), str(row.get("ano","")).split(".")[0],
                     norm(row.get("tipo_documento","")), str(row.get("pagina","")).split(".")[0],
                     norm(row.get("termo_chave","")), norm(row.get("termo_encontrado",""))])

def preencher(arq_preenchido, arq_alvo, arq_saida):
    fp = pd.read_excel(arq_preenchido, sheet_name="codificar")
    al = pd.read_excel(arq_alvo, sheet_name="codificar")
    # garante que as colunas de rotulo aceitem texto (evita erro de dtype no pandas novo)
    for c in ["rotulo_humano_1", "rotulo_humano_2", "observacao_humano"]:
        if c not in al.columns:
            al[c] = ""
        al[c] = al[c].astype(object).where(al[c].notna(), "")
    def _txt(v):
        return "" if (v is None or (isinstance(v, float) and pd.isna(v))) else str(v)
    mapa = {}
    for _, r in fp.iterrows():
        if str(r.get("rotulo_humano_1","")).strip():
            mapa[chave(r)] = (r.get("rotulo_humano_1",""), r.get("rotulo_humano_2",""), r.get("observacao_humano",""))
    n = 0
    for i, r in al.iterrows():
        k = chave(r)
        if k in mapa:
            al.at[i, "rotulo_humano_1"] = _txt(mapa[k][0])
            al.at[i, "rotulo_humano_2"] = _txt(mapa[k][1])
            al.at[i, "observacao_humano"] = _txt(mapa[k][2])
            n += 1
    rot = ["ocorrencia territorial relevante","ocorrencia validada (tecnica)",
           "ocorrencia contextual","ocorrencia administrativa generica / falso positivo"]
    with pd.ExcelWriter(arq_saida, engine="openpyxl") as w:
        al.to_excel(w, sheet_name="codificar", index=False)
        pd.DataFrame({"rotulos_validos": rot}).to_excel(w, sheet_name="instrucoes", index=False)
    print("Preenchidas automaticamente %d de %d linhas (correspondencia exata)." % (n, len(al)))
    print("As linhas sem correspondencia ficaram em branco para codificacao manual.")
    print("Salvo em:", arq_saida)

def remover(arq_param, termos):
    termos_norm = {norm(t) for t in termos}
    wb = openpyxl.load_workbook(arq_param)
    def del_por_coluna(aba, col_idx):
        if aba not in wb.sheetnames: return 0
        ws = wb[aba]; apagar = []
        for r in range(3, ws.max_row+1):
            v = ws.cell(r, col_idx).value
            if v is not None and norm(v) in termos_norm: apagar.append(r)
        for r in reversed(apagar): ws.delete_rows(r, 1)
        return len(apagar)
    n_tb = del_por_coluna("Termos_Busca", 1)
    n_tp = del_por_coluna("Tema_Padrao", 1)
    n_ft = del_por_coluna("Forca_Tecnica", 1)
    # Projetos: tira o termo da lista 'termos_chave_associados' (coluna 4)
    n_pj = 0
    if "Projetos" in wb.sheetnames:
        ws = wb["Projetos"]
        for r in range(3, ws.max_row+1):
            c = ws.cell(r, 4).value
            if not c: continue
            itens = [x for x in str(c).split(";") if norm(x) not in termos_norm]
            novo = "; ".join(s.strip() for s in itens)
            if novo != str(c): ws.cell(r, 4, novo); n_pj += 1
    wb.save(arq_param)
    print("Removidos termos:", ", ".join(termos))
    print("  Termos_Busca: %d linhas | Tema_Padrao: %d | Forca_Tecnica: %d | Projetos ajustados: %d" % (n_tb, n_tp, n_ft, n_pj))
    print("Salvo em:", arq_param)

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print(__doc__); sys.exit(1)
    modo = sys.argv[1].lower()
    if modo == "preencher":
        preencher(sys.argv[2], sys.argv[3], sys.argv[4])
    elif modo == "remover":
        remover(sys.argv[2], sys.argv[3:])
    else:
        sys.exit("Modo invalido: use 'preencher' ou 'remover'.")
