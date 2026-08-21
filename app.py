# -*- coding: utf-8 -*-
"""
Interface grafica para as analises da tese (Instrumentos Urbanisticos / LDO e LOA).

Esta janela nao reimplementa a analise: ela apenas chama, na ordem certa e com os
caminhos escolhidos pelo usuario, os mesmos scripts do repositorio
Tese_Analises_BaixadaFluminense_YasminAlmeida (copiados, com o minimo de ajustes
necessarios, em scripts/instrumentos e scripts/ldo_loa - ver README.md).

Rodar com:  python app.py
"""
import os
import sys
import glob
import json
import shutil
import queue
import threading
import traceback
import subprocess
import unicodedata
import importlib
from pathlib import Path
from datetime import datetime

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext

try:
    import openpyxl
except Exception:
    openpyxl = None

BASE_DIR = Path(__file__).resolve().parent
SCRIPTS_INSTR = BASE_DIR / "scripts" / "instrumentos"
SCRIPTS_LDO = BASE_DIR / "scripts" / "ldo_loa"
DEFAULT_PARAMETROS = BASE_DIR / "parametros" / "Parametros_LDO_LOA.xlsx"

TESSERACT_CANDIDATOS = [
    r"C:\Program Files\Tesseract-OCR\tesseract.exe",
    r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
]

MODULO_PARA_PIP = {
    "fitz": "pymupdf", "PIL": "pillow", "pytesseract": "pytesseract",
    "openpyxl": "openpyxl", "rapidfuzz": "rapidfuzz", "pandas": "pandas",
    "docx": "python-docx", "sklearn": "scikit-learn",
}

# nome do parametro -> (tipo, rotulo em portugues)
PARAM_FIELDS = [
    ("ANO_INICIAL", "int", "Ano inicial da serie (a partir de que ano buscar LDO/LOA)"),
    ("ANO_FINAL", "int", "Ano final da serie"),
    ("TIPOS_DOCUMENTO", "text", "Tipos de documento a buscar, separados por ; (ex.: LDO; LOA)"),
    ("USAR_OCR", "bool", "Usar OCR nas paginas digitalizadas"),
    ("IDIOMA_OCR", "text", "Idioma do OCR (codigo do Tesseract, ex.: por)"),
    ("MIN_CARACTERES_TEXTO_DIGITAL", "int", "Minimo de caracteres para considerar 'texto digital' (abaixo disso, tenta OCR)"),
    ("FORCAR_OCR", "bool", "Forcar OCR mesmo com texto digital, se ele parecer 'lixo'"),
    ("OCR_ZOOM", "int", "Zoom da imagem para OCR (maior = mais nitido e mais lento)"),
    ("JANELA_CONTEXTO", "int", "Tamanho da janela de contexto ao redor de cada termo (caracteres)"),
    ("USAR_FUZZY", "bool", "Usar busca aproximada (fuzzy) para termos corrompidos por OCR"),
    ("FUZZY_LIMIAR", "int", "Limiar de similaridade do fuzzy (0 a 100)"),
    ("FUZZY_SO_EM_PAGINAS_OCR", "bool", "Aplicar fuzzy somente em paginas que passaram por OCR"),
    ("DEDUP_BOILERPLATE", "bool", "Remover ocorrencias duplicadas (mesmo trecho) automaticamente"),
    ("DEDUP_TAM_HASH", "int", "Numero de caracteres do trecho usado para identificar duplicatas"),
    ("EXTRAIR_ORCAMENTO", "bool", "Tentar extrair valores monetarios (R$) proximos a cada ocorrencia"),
]


# ----------------------------------------------------------------------------
# Helpers de sistema de arquivos / dependencias
# ----------------------------------------------------------------------------
def normalizar_simples(s):
    s = unicodedata.normalize("NFKD", s or "")
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s.lower()


TUTORIAL_TESSERACT = (
    "COMO INSTALAR O TESSERACT (necessario para ler PDFs escaneados)\n"
    "\n"
    "O Tesseract e' um PROGRAMA separado do Python -- instalar o pacote 'pytesseract'\n"
    "pelo pip NAO instala o programa. Por isso a analise pode passar na checagem de\n"
    "pacotes e mesmo assim falhar na hora de fazer OCR.\n"
    "\n"
    "1) Baixe o instalador para Windows em:\n"
    "   https://github.com/UB-Mannheim/tesseract/wiki\n"
    "\n"
    "2) Durante a instalacao, em 'Additional language data', marque:\n"
    "   - Portuguese  (para OCR em portugues)\n"
    "   - English     (idioma usado por padrao nesta analise)\n"
    "\n"
    "3) Anote a pasta de instalacao. O padrao e':\n"
    "   C:\\Program Files\\Tesseract-OCR\\tesseract.exe\n"
    "\n"
    "4) Volte a esta tela e clique em 'Detectar e testar'. Se mesmo assim nao achar,\n"
    "   clique em 'Procurar...' e aponte para o tesseract.exe manualmente.\n"
    "\n"
    "OBSERVACAO: o instalador NAO adiciona o Tesseract ao PATH do Windows por padrao.\n"
    "E' normal precisar informar o caminho aqui -- nao significa que a instalacao\n"
    "deu errado."
)


def verificar_tesseract(caminho=""):
    """Confirma que o Tesseract EXECUTA de fato e lista os idiomas instalados.

    Verificar que o arquivo existe, ou que o pacote Python importa, nao basta: o
    programa pode estar corrompido, ser de outra arquitetura, ou nao estar no PATH.
    Devolve (ok: bool, detalhe: str, idiomas: list).
    """
    alvo = (caminho or "").strip().strip('"') or detectar_tesseract() or "tesseract"
    if not alvo:
        return False, "Nao foi encontrado nenhum executavel do Tesseract.", []
    try:
        proc = subprocess.run([alvo, "--version"], capture_output=True, text=True, timeout=60)
    except FileNotFoundError:
        return False, f"O programa nao foi encontrado em: {alvo}", []
    except Exception as e:
        return False, f"Encontrado em '{alvo}', mas nao foi possivel executar: {e}", []
    if proc.returncode != 0:
        return False, f"'{alvo}' respondeu com erro (codigo {proc.returncode}).", []
    linhas = (proc.stdout or proc.stderr or "").strip().splitlines()
    versao = linhas[0] if linhas else "versao desconhecida"
    idiomas = []
    try:
        pl = subprocess.run([alvo, "--list-langs"], capture_output=True, text=True, timeout=60)
        idiomas = [l.strip() for l in (pl.stdout or "").strip().splitlines()[1:] if l.strip()]
    except Exception:
        pass
    return True, f"{versao}\nCaminho: {alvo}", idiomas


def detectar_tesseract():
    for c in TESSERACT_CANDIDATOS:
        if os.path.exists(c):
            return c
    achado = shutil.which("tesseract")
    return achado or ""


def modulos_faltando(nomes):
    faltando = []
    for nome in nomes:
        try:
            importlib.import_module(nome)
        except Exception:
            faltando.append(nome)
    return faltando


def comando_pip(nomes_modulo):
    pacotes = sorted({MODULO_PARA_PIP.get(n, n) for n in nomes_modulo})
    return f"{sys.executable} -m pip install " + " ".join(pacotes)


def contar_pdfs_instrumentos(pasta):
    try:
        return len(glob.glob(os.path.join(pasta, "*.pdf")))
    except Exception:
        return 0


def listar_pastas_municipios(pasta):
    p = Path(pasta)
    if not p.exists() or not p.is_dir():
        return []
    out = []
    for it in sorted(p.iterdir()):
        if it.is_dir():
            n = normalizar_simples(it.name)
            if "loa" in n and "ldo" in n and not n.startswith("saida"):
                out.append(it.name)
    return out


# ----------------------------------------------------------------------------
# Leitura / escrita da aba Parametros_Gerais (sem tocar no resto da planilha)
# ----------------------------------------------------------------------------
def carregar_parametros_gerais(xlsx_path):
    if openpyxl is None:
        raise RuntimeError("O pacote 'openpyxl' nao esta instalado (veja requirements.txt).")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    if "Parametros_Gerais" not in wb.sheetnames:
        raise RuntimeError("A planilha selecionada nao tem a aba 'Parametros_Gerais'.")
    ws = wb["Parametros_Gerais"]
    valores = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row:
            continue
        nome = row[0]
        valor = row[1] if len(row) > 1 else None
        if nome is None or str(nome).strip() == "":
            continue
        valores[str(nome).strip()] = "" if valor is None else str(valor)
    wb.close()
    return valores


def salvar_copia_parametros_editados(xlsx_original, edicoes, pasta_destino):
    if openpyxl is None:
        raise RuntimeError("O pacote 'openpyxl' nao esta instalado (veja requirements.txt).")
    os.makedirs(pasta_destino, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    destino = os.path.join(pasta_destino, f"Parametros_LDO_LOA__editado_{ts}.xlsx")
    shutil.copyfile(xlsx_original, destino)
    wb = openpyxl.load_workbook(destino)
    ws = wb["Parametros_Gerais"]
    for r in range(3, ws.max_row + 1):
        nome = ws.cell(r, 1).value
        if nome is None:
            continue
        nome = str(nome).strip()
        if nome in edicoes:
            ws.cell(r, 2).value = edicoes[nome]
    wb.save(destino)
    return destino


# ----------------------------------------------------------------------------
# Frame com barra de rolagem (usado na lista de parametros da LDO/LOA)
# ----------------------------------------------------------------------------
class FrameRolavel(ttk.Frame):
    def __init__(self, parent, altura=260, **kw):
        super().__init__(parent, **kw)
        self.canvas = tk.Canvas(self, height=altura, highlightthickness=0)
        self.scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.interior = ttk.Frame(self.canvas)
        self.interior.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")),
        )
        self.canvas_window = self.canvas.create_window((0, 0), window=self.interior, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)
        self.canvas.bind("<Configure>", self._resize_interior)
        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel, add="+")
        # Este widget e recriado toda vez que o usuario revisita o Passo 3; sem isso, cada
        # visita deixaria um bind_all "morto" apontando para um canvas ja destruido.
        self.bind("<Destroy>", self._ao_destruir)

    def _ao_destruir(self, event):
        if event.widget is self:
            try:
                self.canvas.unbind_all("<MouseWheel>")
            except Exception:
                pass

    def _resize_interior(self, event):
        self.canvas.itemconfig(self.canvas_window, width=event.width)

    def _on_mousewheel(self, event):
        # so rola se o mouse estiver sobre este widget
        try:
            widget = self.canvas.winfo_containing(event.x_root, event.y_root)
            if widget is None:
                return
            if str(widget).startswith(str(self.canvas)):
                self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        except tk.TclError:
            pass


# ----------------------------------------------------------------------------
# Aplicativo principal
# ----------------------------------------------------------------------------
class App:
    TITULOS = ["Tipo de analise", "Pasta de origem", "Parametros", "Revisar e rodar"]

    def __init__(self, root):
        self.root = root
        self.root.title("Analises da tese - Baixada Fluminense Historica")
        self.root.geometry("980x720")
        self.root.minsize(860, 620)

        # ---- estado ----
        self.tipo = tk.StringVar(value="instrumentos")
        self.pasta_origem = tk.StringVar()
        self.parametros_xlsx = tk.StringVar(value=str(DEFAULT_PARAMETROS) if DEFAULT_PARAMETROS.exists() else "")
        self.pasta_saida = tk.StringVar()
        self.arquivo_saida_ldo = tk.StringVar()
        self.tesseract_path = tk.StringVar(value=detectar_tesseract())
        self.ocr_instrumentos = tk.StringVar(value="Sim")
        self.ocr_lang_instrumentos = tk.StringVar(value="eng")

        self.ldo_param_vars = {}       # nome -> tk.StringVar (widget)
        self.ldo_param_original = {}   # nome -> valor original (string)

        self.cancel_flag = threading.Event()
        self.exec_thread = None
        self.current_proc = None
        self.log_queue = queue.Queue()
        self.resultado_path = None
        self.resultado_dir = None

        self.step = 0

        self._montar_layout()
        self._render_step()

    # ------------------------------------------------------------------
    # Layout geral (cabecalho + area de conteudo + barra de navegacao)
    # ------------------------------------------------------------------
    def _montar_layout(self):
        menubar = tk.Menu(self.root)
        ferramentas = tk.Menu(menubar, tearoff=0)
        ferramentas.add_command(label="Ferramentas avancadas (opcionais)...", command=self._abrir_avancado)
        menubar.add_cascade(label="Ferramentas", menu=ferramentas)
        self.root.config(menu=menubar)

        topo = ttk.Frame(self.root, padding=(16, 12, 16, 4))
        topo.pack(fill="x")
        self.lbl_header = ttk.Label(topo, text="", font=("Segoe UI", 13, "bold"))
        self.lbl_header.pack(anchor="w")
        self.lbl_sub = ttk.Label(topo, text="", foreground="#555555")
        self.lbl_sub.pack(anchor="w")

        ttk.Separator(self.root).pack(fill="x", padx=16)

        self.content = ttk.Frame(self.root, padding=16)
        self.content.pack(fill="both", expand=True)

        ttk.Separator(self.root).pack(fill="x", padx=16)

        nav = ttk.Frame(self.root, padding=(16, 10))
        nav.pack(fill="x")
        self.btn_voltar = ttk.Button(nav, text="< Voltar", command=self._voltar)
        self.btn_voltar.pack(side="left")
        self.btn_avancar = ttk.Button(nav, text="Avancar >", command=self._avancar)
        self.btn_avancar.pack(side="right")

    def _limpar_content(self):
        for w in self.content.winfo_children():
            w.destroy()

    def _render_step(self):
        self._limpar_content()
        self.lbl_header.config(text=f"Passo {self.step + 1} de 4 - {self.TITULOS[self.step]}")
        subtitulos = {
            0: "Escolha qual das duas analises da tese voce quer rodar.",
            1: "Escolha a pasta com os documentos de origem.",
            2: "Escolha a planilha de parametros, a pasta/arquivo de saida e ajuste as opcoes.",
            3: "Confira o resumo e clique em Rodar. O andamento aparece no log abaixo.",
        }
        self.lbl_sub.config(text=subtitulos[self.step])

        if self.step == 0:
            self._passo_tipo()
        elif self.step == 1:
            self._passo_pasta()
        elif self.step == 2:
            self._passo_parametros()
        elif self.step == 3:
            self._passo_revisar()

        self.btn_voltar.config(state=("disabled" if self.step == 0 else "normal"))
        if self.step == 3:
            self.btn_avancar.pack_forget()
        else:
            self.btn_avancar.pack(side="right")
            self.btn_avancar.config(text="Avancar >")

    def _voltar(self):
        if self.step > 0:
            self.step -= 1
            self._render_step()

    def _avancar(self):
        if not self._validar_passo_atual():
            return
        if self.step < 3:
            self.step += 1
            self._render_step()

    def _validar_passo_atual(self):
        if self.step == 0:
            if self.tipo.get() not in ("instrumentos", "ldo_loa"):
                messagebox.showerror("Falta escolher", "Escolha um tipo de analise para continuar.")
                return False
            return True
        if self.step == 1:
            pasta = self.pasta_origem.get().strip()
            if not pasta:
                messagebox.showerror("Falta a pasta", "Escolha a pasta de origem dos documentos.")
                return False
            if not os.path.isdir(pasta):
                messagebox.showerror("Pasta invalida", "A pasta escolhida nao existe.")
                return False
            return True
        if self.step == 2:
            return self._validar_parametros()
        return True

    def _validar_parametros(self):
        params = self.parametros_xlsx.get().strip()
        if not params or not os.path.isfile(params):
            messagebox.showerror("Falta a planilha", "Escolha o arquivo Parametros_LDO_LOA.xlsx.")
            return False
        if self.tipo.get() == "instrumentos":
            if not self.pasta_saida.get().strip():
                messagebox.showerror("Falta a pasta de saida", "Escolha a pasta onde os resultados serao gravados.")
                return False
        else:
            if not self.arquivo_saida_ldo.get().strip():
                messagebox.showerror("Falta o arquivo de saida", "Escolha onde o Resultados_LDO_LOA.xlsx sera salvo.")
                return False
        return True

    # ------------------------------------------------------------------
    # Passo 0 - tipo de analise
    # ------------------------------------------------------------------
    def _passo_tipo(self):
        c = self.content

        def bloco(valor, titulo, descricao):
            f = ttk.Frame(c, relief="groove", padding=14)
            f.pack(fill="x", pady=8)
            ttk.Radiobutton(f, text=titulo, value=valor, variable=self.tipo,
                             style="Titulo.TRadiobutton").pack(anchor="w")
            ttk.Label(f, text=descricao, wraplength=820, justify="left",
                      foreground="#444444").pack(anchor="w", pady=(4, 0))

        style = ttk.Style()
        style.configure("Titulo.TRadiobutton", font=("Segoe UI", 11, "bold"))

        bloco(
            "instrumentos",
            "Instrumentos Urbanisticos municipais (item 3.2 da tese)",
            "Le os PDFs de CTM, Planta Generica de Valores, Plano Diretor, Plano de Mobilidade, "
            "Plano de Saneamento e Plano Diretor de TI dos 8 municipios da Baixada Fluminense "
            "Historica; mede o Grau de Integracao Municipal da geoinformacao.",
        )
        bloco(
            "ldo_loa",
            "LDO e LOA municipais (item 3.3 da tese)",
            "Le os PDFs das Leis de Diretrizes Orcamentarias e Leis Orcamentarias Anuais, "
            "organizados em pastas por municipio; mede o Grau de Integracao Orcamentaria da "
            "geoinformacao.",
        )

    # ------------------------------------------------------------------
    # Passo 1 - pasta de origem
    # ------------------------------------------------------------------
    def _passo_pasta(self):
        c = self.content
        if self.tipo.get() == "instrumentos":
            instrucao = (
                "Escolha a pasta que contem os PDFs dos instrumentos urbanisticos, um arquivo por "
                "instrumento e municipio (ex.: 'CTM - Japeri.pdf', 'Plano Diretor - Duque de Caxias.pdf')."
            )
        else:
            instrucao = (
                "Escolha a pasta-raiz que contem uma subpasta por municipio, no padrao "
                "'<MUNICIPIO> LOA E LDO' (ex.: 'BELFORD ROXO LOA E LDO'), cada uma com os PDFs de "
                "LDO e LOA daquele municipio."
            )
        ttk.Label(c, text=instrucao, wraplength=880, justify="left").pack(anchor="w", pady=(0, 12))

        linha = ttk.Frame(c)
        linha.pack(fill="x")
        ttk.Entry(linha, textvariable=self.pasta_origem).pack(side="left", fill="x", expand=True)
        ttk.Button(linha, text="Procurar...", command=self._escolher_pasta_origem).pack(side="left", padx=(8, 0))

        self.lbl_scan = ttk.Label(c, text="", foreground="#2e7d32")
        self.lbl_scan.pack(anchor="w", pady=(10, 0))
        ttk.Button(c, text="Verificar pasta", command=self._verificar_pasta_origem).pack(anchor="w", pady=(6, 0))

        if self.pasta_origem.get():
            self._verificar_pasta_origem()

    def _escolher_pasta_origem(self):
        pasta = filedialog.askdirectory(title="Escolha a pasta de origem")
        if pasta:
            self.pasta_origem.set(pasta)
            self._verificar_pasta_origem()

    def _verificar_pasta_origem(self):
        pasta = self.pasta_origem.get().strip()
        if not pasta or not os.path.isdir(pasta):
            self.lbl_scan.config(text="Pasta nao encontrada.", foreground="#c0392b")
            return
        if self.tipo.get() == "instrumentos":
            n = contar_pdfs_instrumentos(pasta)
            if n == 0:
                self.lbl_scan.config(text="Nenhum PDF encontrado diretamente nesta pasta.", foreground="#c0392b")
            else:
                self.lbl_scan.config(text=f"{n} PDF(s) encontrado(s) nesta pasta.", foreground="#2e7d32")
        else:
            municipios = listar_pastas_municipios(pasta)
            if not municipios:
                self.lbl_scan.config(
                    text="Nenhuma subpasta '<MUNICIPIO> LOA E LDO' encontrada.", foreground="#c0392b")
            else:
                self.lbl_scan.config(
                    text=f"{len(municipios)} pasta(s) de municipio encontrada(s): " + ", ".join(municipios),
                    foreground="#2e7d32")

    # ------------------------------------------------------------------
    # Passo 2 - parametros
    # ------------------------------------------------------------------
    def _passo_parametros(self):
        c = self.content
        wrapper = FrameRolavel(c, altura=460)
        wrapper.pack(fill="both", expand=True)
        f = wrapper.interior

        # ---- planilha de parametros (comum aos dois eixos) ----
        grp = ttk.LabelFrame(f, text="Planilha de parametros (vocabulario, pesos, rubrica)", padding=10)
        grp.pack(fill="x", pady=(0, 10))
        linha = ttk.Frame(grp)
        linha.pack(fill="x")
        ttk.Entry(linha, textvariable=self.parametros_xlsx).pack(side="left", fill="x", expand=True)
        ttk.Button(linha, text="Procurar...", command=self._escolher_parametros).pack(side="left", padx=(8, 0))
        ttk.Label(grp, text="Parametros_LDO_LOA.xlsx - Anexos A/B/C da tese. Feche o arquivo no Excel antes de rodar.",
                  foreground="#666666").pack(anchor="w", pady=(4, 0))

        # ---- tesseract ----
        grp2 = ttk.LabelFrame(f, text="OCR (Tesseract)", padding=10)
        grp2.pack(fill="x", pady=(0, 10))
        linha2 = ttk.Frame(grp2)
        linha2.pack(fill="x")
        ttk.Entry(linha2, textvariable=self.tesseract_path).pack(side="left", fill="x", expand=True)
        ttk.Button(linha2, text="Procurar...", command=self._escolher_tesseract).pack(side="left", padx=(8, 0))
        ttk.Button(linha2, text="Detectar automaticamente", command=self._detectar_tesseract_click).pack(
            side="left", padx=(8, 0))
        ttk.Label(grp2, text="Caminho do tesseract.exe. Deixe em branco se ja estiver no PATH do sistema.",
                  foreground="#666666").pack(anchor="w", pady=(4, 0))

        # ---- saida ----
        grp3 = ttk.LabelFrame(f, text="Saida", padding=10)
        grp3.pack(fill="x", pady=(0, 10))
        if self.tipo.get() == "instrumentos":
            linha3 = ttk.Frame(grp3)
            linha3.pack(fill="x")
            ttk.Label(linha3, text="Pasta de saida:").pack(side="left")
            ttk.Entry(linha3, textvariable=self.pasta_saida).pack(side="left", fill="x", expand=True, padx=8)
            ttk.Button(linha3, text="Procurar...", command=self._escolher_pasta_saida).pack(side="left")
            ttk.Label(grp3,
                      text="Sera criada, se nao existir. Vai receber resultados.json, ocorrencias.json, "
                           "analise_final.json e o Excel final (Analise_Institucionalizacao_Geoinformacao.xlsx).",
                      foreground="#666666", wraplength=820, justify="left").pack(anchor="w", pady=(4, 0))
        else:
            linha3 = ttk.Frame(grp3)
            linha3.pack(fill="x")
            ttk.Label(linha3, text="Pasta de trabalho:").pack(side="left")
            ttk.Entry(linha3, textvariable=self.pasta_saida).pack(side="left", fill="x", expand=True, padx=8)
            ttk.Button(linha3, text="Procurar...", command=self._escolher_pasta_saida_ldo).pack(side="left")
            linha4 = ttk.Frame(grp3)
            linha4.pack(fill="x", pady=(6, 0))
            ttk.Label(linha4, text="Arquivo Excel de saida:").pack(side="left")
            ttk.Entry(linha4, textvariable=self.arquivo_saida_ldo).pack(side="left", fill="x", expand=True, padx=8)
            ttk.Button(linha4, text="Procurar...", command=self._escolher_arquivo_saida_ldo).pack(side="left")

        # ---- opcoes especificas ----
        if self.tipo.get() == "instrumentos":
            grp4 = ttk.LabelFrame(f, text="Opcoes da analise de Instrumentos", padding=10)
            grp4.pack(fill="x", pady=(0, 10))
            l1 = ttk.Frame(grp4)
            l1.pack(fill="x", pady=2)
            ttk.Label(l1, text="Usar OCR nas paginas digitalizadas:", width=42).pack(side="left")
            ttk.Combobox(l1, textvariable=self.ocr_instrumentos, values=["Sim", "Nao"],
                         state="readonly", width=10).pack(side="left")
            l2 = ttk.Frame(grp4)
            l2.pack(fill="x", pady=2)
            ttk.Label(l2, text="Idioma do OCR:", width=42).pack(side="left")
            ttk.Combobox(l2, textvariable=self.ocr_lang_instrumentos, values=["eng", "por"],
                         state="readonly", width=10).pack(side="left")
            ttk.Label(grp4,
                      text="'eng' e o valor original do script (usado nos resultados publicados no item 4.3 da "
                           "tese) - possivelmente um erro de digitacao, ja que os documentos sao em portugues. "
                           "Troque para 'por' apenas se quiser reprocessar/testar; isso muda os resultados do OCR.",
                      foreground="#a15c00", wraplength=820, justify="left").pack(anchor="w", pady=(4, 0))
        else:
            self._bloco_parametros_ldo(f)

    def _bloco_parametros_ldo(self, f):
        grp = ttk.LabelFrame(f, text="Parametros gerais da analise (aba Parametros_Gerais)", padding=10)
        grp.pack(fill="both", expand=True, pady=(0, 10))

        topo = ttk.Frame(grp)
        topo.pack(fill="x", pady=(0, 8))
        ttk.Button(topo, text="Carregar parametros da planilha selecionada",
                   command=self._carregar_parametros_ldo).pack(side="left")
        self.lbl_ldo_status = ttk.Label(topo, text="", foreground="#666666")
        self.lbl_ldo_status.pack(side="left", padx=10)

        self.frame_campos_ldo = ttk.Frame(grp)
        self.frame_campos_ldo.pack(fill="both", expand=True)

        if self.ldo_param_vars:
            self._desenhar_campos_ldo()
        else:
            ttk.Label(self.frame_campos_ldo,
                      text="Clique em 'Carregar parametros da planilha selecionada' para ver e editar os "
                           "14 parametros gerais usados pela analise de LDO/LOA.",
                      foreground="#666666", wraplength=800, justify="left").pack(anchor="w")

    def _carregar_parametros_ldo(self):
        caminho = self.parametros_xlsx.get().strip()
        if not caminho or not os.path.isfile(caminho):
            messagebox.showerror("Planilha nao encontrada", "Escolha primeiro um arquivo Parametros_LDO_LOA.xlsx valido.")
            return
        try:
            valores = carregar_parametros_gerais(caminho)
        except Exception as e:
            messagebox.showerror("Erro ao ler a planilha", str(e))
            return

        self.ldo_param_original = {}
        self.ldo_param_vars = {}
        for nome, tipo_campo, _rotulo in PARAM_FIELDS:
            bruto = valores.get(nome, "")
            if tipo_campo == "bool":
                bruto = "Sim" if str(bruto).strip().lower() in ("sim", "s", "true", "1", "verdadeiro") else "Nao"
            elif tipo_campo == "int" and str(bruto).strip() != "":
                # celulas numericas do Excel podem vir como "2015.0"; normaliza para "2015"
                # para exibir limpo e para nao gravar de volta um texto que o script original
                # (que espera int(valor)) nao consiga converter.
                try:
                    bruto = str(int(float(bruto)))
                except ValueError:
                    pass
            self.ldo_param_original[nome] = bruto
            self.ldo_param_vars[nome] = tk.StringVar(value=bruto)

        faltando = [n for n, _, _ in PARAM_FIELDS if n not in valores]
        if faltando:
            self.lbl_ldo_status.config(
                text=f"Carregado. Aviso: nao encontrado na planilha: {', '.join(faltando)} (ficou em branco).",
                foreground="#a15c00")
        else:
            self.lbl_ldo_status.config(text="Parametros carregados com sucesso.", foreground="#2e7d32")

        self._desenhar_campos_ldo()

    def _desenhar_campos_ldo(self):
        for w in self.frame_campos_ldo.winfo_children():
            w.destroy()
        for nome, tipo_campo, rotulo in PARAM_FIELDS:
            var = self.ldo_param_vars[nome]
            linha = ttk.Frame(self.frame_campos_ldo)
            linha.pack(fill="x", pady=3)
            ttk.Label(linha, text=nome, width=26, font=("Segoe UI", 9, "bold")).pack(side="left")
            if tipo_campo == "bool":
                ttk.Combobox(linha, textvariable=var, values=["Sim", "Nao"], state="readonly", width=10).pack(
                    side="left")
            elif tipo_campo == "int":
                ttk.Spinbox(linha, textvariable=var, from_=0, to=99999, width=10).pack(side="left")
            else:
                ttk.Entry(linha, textvariable=var, width=20).pack(side="left")
            ttk.Label(linha, text=rotulo, foreground="#666666", wraplength=520, justify="left").pack(
                side="left", padx=(10, 0))

    def _escolher_parametros(self):
        caminho = filedialog.askopenfilename(
            title="Escolha o arquivo Parametros_LDO_LOA.xlsx",
            filetypes=[("Planilha Excel", "*.xlsx")])
        if caminho:
            self.parametros_xlsx.set(caminho)
            self.ldo_param_vars = {}
            if self.tipo.get() == "ldo_loa":
                self._render_step()

    def _escolher_tesseract(self):
        caminho = filedialog.askopenfilename(
            title="Escolha o tesseract.exe", filetypes=[("Executavel", "*.exe"), ("Todos", "*.*")])
        if caminho:
            self.tesseract_path.set(caminho)

    def _detectar_tesseract_click(self):
        achado = detectar_tesseract()
        if achado:
            self.tesseract_path.set(achado)
        ok, detalhe, idiomas = verificar_tesseract(self.tesseract_path.get().strip())
        if ok:
            lista = ", ".join(idiomas) if idiomas else "(nao foi possivel listar)"
            messagebox.showinfo(
                "Tesseract funcionando",
                f"O Tesseract foi encontrado e executou corretamente.\n\n{detalhe}\n\n"
                f"Idiomas de OCR instalados: {lista}")
        else:
            messagebox.showwarning(
                "Tesseract nao esta funcionando",
                f"{detalhe}\n\n{TUTORIAL_TESSERACT}")

    def _escolher_pasta_saida(self):
        pasta = filedialog.askdirectory(title="Escolha a pasta de saida")
        if pasta:
            self.pasta_saida.set(pasta)

    def _escolher_pasta_saida_ldo(self):
        pasta = filedialog.askdirectory(title="Escolha a pasta de trabalho")
        if pasta:
            self.pasta_saida.set(pasta)
            if not self.arquivo_saida_ldo.get().strip():
                self.arquivo_saida_ldo.set(os.path.join(pasta, "Resultados_LDO_LOA.xlsx"))

    def _escolher_arquivo_saida_ldo(self):
        caminho = filedialog.asksaveasfilename(
            title="Onde salvar o Excel de resultados", defaultextension=".xlsx",
            initialfile="Resultados_LDO_LOA.xlsx", filetypes=[("Planilha Excel", "*.xlsx")])
        if caminho:
            self.arquivo_saida_ldo.set(caminho)

    # ------------------------------------------------------------------
    # Passo 3 - revisar e rodar
    # ------------------------------------------------------------------
    def _passo_revisar(self):
        c = self.content

        resumo = ttk.LabelFrame(c, text="Resumo", padding=10)
        resumo.pack(fill="x")

        linhas = []
        if self.tipo.get() == "instrumentos":
            linhas.append(("Analise", "Instrumentos Urbanisticos (item 3.2)"))
            linhas.append(("Pasta de origem (PDFs)", self.pasta_origem.get()))
            linhas.append(("Planilha de parametros", self.parametros_xlsx.get()))
            linhas.append(("Pasta de saida", self.pasta_saida.get()))
            linhas.append(("OCR", f"{self.ocr_instrumentos.get()} (idioma: {self.ocr_lang_instrumentos.get()})"))
        else:
            linhas.append(("Analise", "LDO e LOA (item 3.3)"))
            linhas.append(("Pasta de origem (raiz dos municipios)", self.pasta_origem.get()))
            linhas.append(("Planilha de parametros", self.parametros_xlsx.get()))
            linhas.append(("Arquivo de saida", self.arquivo_saida_ldo.get()))
            edits = self._coletar_edicoes_ldo()
            if edits:
                linhas.append(("Parametros gerais alterados", ", ".join(f"{k}={v}" for k, v in edits.items())))
            else:
                linhas.append(("Parametros gerais", "sem alteracoes (usa a planilha original)"))
        linhas.append(("Tesseract", self.tesseract_path.get() or "(procurar no PATH do sistema)"))

        for rotulo, valor in linhas:
            l = ttk.Frame(resumo)
            l.pack(fill="x", pady=1)
            ttk.Label(l, text=rotulo + ":", width=32, anchor="w").pack(side="left")
            ttk.Label(l, text=valor or "-", foreground="#333333", wraplength=560, justify="left").pack(
                side="left", fill="x", expand=True)

        botoes = ttk.Frame(c)
        botoes.pack(fill="x", pady=10)
        self.btn_rodar = ttk.Button(botoes, text="Rodar", command=self._on_rodar)
        self.btn_rodar.pack(side="left")
        self.btn_cancelar = ttk.Button(botoes, text="Cancelar", command=self._on_cancelar, state="disabled")
        self.btn_cancelar.pack(side="left", padx=8)
        self.btn_abrir_pasta = ttk.Button(botoes, text="Abrir pasta de resultados", command=self._abrir_pasta_resultado,
                                           state="disabled")
        self.btn_abrir_pasta.pack(side="left", padx=8)
        self.btn_abrir_arquivo = ttk.Button(botoes, text="Abrir arquivo de resultados", command=self._abrir_arquivo_resultado,
                                             state="disabled")
        self.btn_abrir_arquivo.pack(side="left", padx=8)

        self.lbl_progresso = ttk.Label(c, text="Pronto para rodar.")
        self.lbl_progresso.pack(anchor="w", pady=(6, 2))
        self.progress = ttk.Progressbar(c, mode="determinate", maximum=1, value=0)
        self.progress.pack(fill="x")

        ttk.Label(c, text="Log:").pack(anchor="w", pady=(10, 2))
        self.txt_log = scrolledtext.ScrolledText(c, height=16, state="disabled", font=("Consolas", 9))
        self.txt_log.pack(fill="both", expand=True)

    # ------------------------------------------------------------------
    # Execucao
    # ------------------------------------------------------------------
    def _coletar_edicoes_ldo(self):
        tipos = {n: t for n, t, _r in PARAM_FIELDS}
        edits = {}
        for nome, var in self.ldo_param_vars.items():
            novo = var.get()
            original = self.ldo_param_original.get(nome, "")
            if str(novo).strip() != str(original).strip():
                if tipos.get(nome) == "int":
                    try:
                        novo = int(float(novo))
                    except ValueError:
                        pass
                edits[nome] = novo
        return edits

    def _checar_dependencias(self):
        obrigatorios = ["fitz", "openpyxl"]
        if self.tipo.get() == "ldo_loa":
            obrigatorios.append("pandas")
        recomendados = ["pytesseract", "PIL", "rapidfuzz"]

        falt_obrig = modulos_faltando(obrigatorios)
        if falt_obrig:
            messagebox.showerror(
                "Faltam pacotes obrigatorios",
                "Estes pacotes Python nao foram encontrados e sao obrigatorios:\n\n"
                f"{', '.join(falt_obrig)}\n\nInstale rodando no terminal:\n{comando_pip(falt_obrig)}")
            return False

        falt_rec = modulos_faltando(recomendados)
        if falt_rec:
            prosseguir = messagebox.askyesno(
                "Pacotes recomendados ausentes",
                "Estes pacotes nao foram encontrados: " + ", ".join(falt_rec) + ".\n"
                "Sem eles o OCR e/ou a busca aproximada (fuzzy) ficam desativados, mas a analise ainda roda.\n\n"
                f"Instale com:\n{comando_pip(falt_rec)}\n\nContinuar mesmo assim?")
            if not prosseguir:
                return False

        # ---- verificacao do PROGRAMA Tesseract (nao so' do pacote Python) ----
        # Checar que o modulo 'pytesseract' importa NAO garante que o OCR va' funcionar:
        # o Tesseract e' um programa a parte. Sem este teste, a analise so' descobria o
        # problema depois de varios minutos de processamento, ja' dentro da Etapa 2.
        if self._ocr_solicitado():
            ok, detalhe, idiomas = verificar_tesseract(self.tesseract_path.get().strip())
            if not ok:
                prosseguir = messagebox.askyesno(
                    "Tesseract nao esta funcionando",
                    f"Voce pediu para usar OCR, mas o Tesseract nao pode ser executado.\n\n"
                    f"{detalhe}\n\n{TUTORIAL_TESSERACT}\n\n"
                    "Deseja rodar mesmo assim, SEM OCR? Os PDFs escaneados serao lidos apenas\n"
                    "pelo texto ja' embutido no arquivo, o que reduz a cobertura e pode mudar\n"
                    "os resultados em relacao aos publicados na tese.")
                if not prosseguir:
                    return False
            else:
                idioma = self._idioma_ocr_escolhido()
                pedidos = [p.strip() for p in idioma.split("+") if p.strip()]
                faltam = [p for p in pedidos if idiomas and p not in idiomas]
                if faltam:
                    prosseguir = messagebox.askyesno(
                        "Idioma de OCR nao instalado",
                        f"O Tesseract esta funcionando, mas o(s) idioma(s) "
                        f"{', '.join(faltam)} nao esta(o) instalado(s).\n\n"
                        f"Idiomas disponiveis: {', '.join(idiomas)}\n\n"
                        "Para instalar, rode o instalador do Tesseract novamente e marque o\n"
                        "idioma em 'Additional language data'.\n\n"
                        "Rodar mesmo assim? O OCR provavelmente vai falhar nas paginas escaneadas.")
                    if not prosseguir:
                        return False
        return True

    def _ocr_solicitado(self):
        """A analise escolhida vai tentar usar OCR?"""
        if self.tipo.get() == "instrumentos":
            return self.ocr_instrumentos.get() == "Sim"
        var = self.ldo_param_vars.get("USAR_OCR")
        valor = var.get() if var is not None else self.ldo_param_original.get("USAR_OCR", "")
        return str(valor).strip().lower() in ("1", "sim", "s", "true", "verdadeiro", "yes", "y")

    def _idioma_ocr_escolhido(self):
        """Codigo de idioma que sera' passado ao Tesseract (ex.: 'eng', 'por')."""
        if self.tipo.get() == "instrumentos":
            return (self.ocr_lang_instrumentos.get() or "eng").strip()
        var = self.ldo_param_vars.get("IDIOMA_OCR")
        valor = var.get() if var is not None else self.ldo_param_original.get("IDIOMA_OCR", "por")
        return (str(valor) or "por").strip()

    def _on_rodar(self):
        if not self._validar_parametros():
            return
        if not self._checar_dependencias():
            return
        self.cancel_flag.clear()
        self.resultado_path = None
        self.resultado_dir = None
        self.txt_log.config(state="normal")
        self.txt_log.delete("1.0", "end")
        self.txt_log.config(state="disabled")
        self.btn_rodar.config(state="disabled")
        self.btn_cancelar.config(state="normal")
        self.btn_abrir_pasta.config(state="disabled")
        self.btn_abrir_arquivo.config(state="disabled")
        self.btn_voltar.config(state="disabled")

        self.exec_thread = threading.Thread(target=self._executar, daemon=True)
        self.exec_thread.start()
        self.root.after(150, self._poll_queue)

    def _on_cancelar(self):
        self.cancel_flag.set()
        if self.current_proc is not None:
            try:
                self.current_proc.terminate()
            except Exception:
                pass
        self.btn_cancelar.config(state="disabled")

    def _log(self, msg):
        self.log_queue.put(("log", msg))

    def _set_progress(self, idx, total, desc):
        self.log_queue.put(("progress", (idx, total, desc)))

    def _poll_queue(self):
        try:
            while True:
                kind, payload = self.log_queue.get_nowait()
                if kind == "log":
                    self.txt_log.config(state="normal")
                    self.txt_log.insert("end", payload + "\n")
                    self.txt_log.see("end")
                    self.txt_log.config(state="disabled")
                elif kind == "progress":
                    idx, total, desc = payload
                    self.progress.config(maximum=max(total, 1), value=idx)
                    self.lbl_progresso.config(text=desc)
                elif kind == "done":
                    ok, erro = payload
                    self._on_finish(ok, erro)
        except queue.Empty:
            pass
        if self.exec_thread is not None and self.exec_thread.is_alive():
            self.root.after(150, self._poll_queue)

    def _on_finish(self, ok, erro):
        self.btn_rodar.config(state="normal")
        self.btn_cancelar.config(state="disabled")
        self.btn_voltar.config(state="normal")
        if ok:
            self.lbl_progresso.config(text="Concluido.")
            if self.resultado_dir and os.path.isdir(self.resultado_dir):
                self.btn_abrir_pasta.config(state="normal")
            if self.resultado_path and os.path.isfile(self.resultado_path):
                self.btn_abrir_arquivo.config(state="normal")
            messagebox.showinfo("Concluido", "Analise concluida com sucesso.")
        else:
            self.lbl_progresso.config(text="Terminou com erro.")
            messagebox.showerror("Erro", erro or "A analise terminou com erro. Veja o log para detalhes.")

    def _run_subprocess(self, cmd, cwd=None, env=None):
        self._log("$ " + " ".join(str(x) for x in cmd) + (f"   (pasta: {cwd})" if cwd else ""))
        try:
            proc = subprocess.Popen(
                cmd, cwd=cwd, env=env, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                text=True, encoding="utf-8", errors="replace", bufsize=1)
        except FileNotFoundError as e:
            raise RuntimeError(f"Nao foi possivel iniciar o processo: {e}")
        self.current_proc = proc
        for linha in proc.stdout:
            if self.cancel_flag.is_set():
                proc.terminate()
                self._log("[cancelado pelo usuario]")
                break
            self._log(linha.rstrip("\n"))
        proc.wait()
        self.current_proc = None
        if self.cancel_flag.is_set():
            raise RuntimeError("Execucao cancelada pelo usuario.")
        if proc.returncode != 0:
            raise RuntimeError(f"O script terminou com codigo de erro {proc.returncode}. Veja o log acima.")

    def _executar(self):
        try:
            env = os.environ.copy()
            tess = self.tesseract_path.get().strip()
            if tess:
                env["TESSERACT_CMD"] = tess
                env["LDO_TESSERACT_PATH"] = tess
            if self.tipo.get() == "instrumentos":
                self._executar_instrumentos(env)
            else:
                self._executar_ldo(env)
            self.log_queue.put(("done", (True, None)))
        except Exception as e:
            self._log("ERRO: " + str(e))
            self._log(traceback.format_exc())
            self.log_queue.put(("done", (False, str(e))))

    def _executar_instrumentos(self, env):
        pasta_pdfs = os.path.abspath(self.pasta_origem.get().strip())
        parametros = os.path.abspath(self.parametros_xlsx.get().strip())
        pasta_saida = os.path.abspath(self.pasta_saida.get().strip())
        os.makedirs(pasta_saida, exist_ok=True)
        saida_bruta = os.path.join(pasta_saida, "saida")
        os.makedirs(saida_bruta, exist_ok=True)

        env_analise = dict(env)
        env_analise["OCR_LANG"] = self.ocr_lang_instrumentos.get()

        self._set_progress(0, 3, "Etapa 1/3 - extraindo texto e classificando (analise_instrumentos.py)")
        self._run_subprocess(
            [sys.executable, str(SCRIPTS_INSTR / "analise_instrumentos.py"),
             "--pasta", pasta_pdfs, "--parametros", parametros, "--saida", saida_bruta,
             "--ocr", ("sim" if self.ocr_instrumentos.get() == "Sim" else "nao")],
            cwd=None, env=env_analise)

        shutil.copyfile(os.path.join(saida_bruta, "resultados.json"), os.path.join(pasta_saida, "resultados.json"))
        shutil.copyfile(os.path.join(saida_bruta, "ocorrencias.json"), os.path.join(pasta_saida, "ocorrencias.json"))

        env_enhance = dict(env)
        env_enhance["PASTA_INSTRUMENTOS"] = pasta_pdfs
        env_enhance["PARAMETROS_XLSX"] = parametros
        self._set_progress(1, 3, "Etapa 2/3 - enriquecimento e datacao (enhance.py)")
        self._run_subprocess([sys.executable, str(SCRIPTS_INSTR / "enhance.py")], cwd=pasta_saida, env=env_enhance)

        self._set_progress(2, 3, "Etapa 3/3 - montagem da planilha final (build_final.py)")
        self._run_subprocess([sys.executable, str(SCRIPTS_INSTR / "build_final.py")], cwd=pasta_saida, env=env)

        self._set_progress(3, 3, "Concluido.")
        self.resultado_path = os.path.join(pasta_saida, "Analise_Institucionalizacao_Geoinformacao.xlsx")
        self.resultado_dir = pasta_saida

    def _executar_ldo(self, env):
        pasta_raiz = os.path.abspath(self.pasta_origem.get().strip())
        arquivo_saida = os.path.abspath(self.arquivo_saida_ldo.get().strip())
        pasta_destino = os.path.dirname(arquivo_saida) or os.path.abspath(self.pasta_saida.get().strip() or ".")
        os.makedirs(pasta_destino, exist_ok=True)

        parametros_originais = os.path.abspath(self.parametros_xlsx.get().strip())
        edits = self._coletar_edicoes_ldo()
        if edits:
            parametros_finais = salvar_copia_parametros_editados(parametros_originais, edits, pasta_destino)
            self._log(f"Parametros gerais alterados - usando copia: {parametros_finais}")
        else:
            parametros_finais = parametros_originais

        env_run = dict(env)
        env_run["LDO_ARQ_PARAMETROS"] = parametros_finais
        env_run["LDO_PASTA_RAIZ"] = pasta_raiz
        env_run["LDO_ARQ_SAIDA"] = arquivo_saida

        self._set_progress(0, 1, "Rodando analise LDO/LOA (pode demorar bastante com OCR em documentos grandes)")
        self._run_subprocess([sys.executable, str(SCRIPTS_LDO / "analise_ldo_loa.py")], cwd=None, env=env_run)

        self._set_progress(1, 1, "Concluido.")
        self.resultado_path = arquivo_saida
        self.resultado_dir = pasta_destino

    def _abrir_pasta_resultado(self):
        if self.resultado_dir and os.path.isdir(self.resultado_dir):
            os.startfile(self.resultado_dir)

    def _abrir_arquivo_resultado(self):
        if self.resultado_path and os.path.isfile(self.resultado_path):
            os.startfile(self.resultado_path)

    # ------------------------------------------------------------------
    # Ferramentas avancadas (opcionais)
    # ------------------------------------------------------------------
    def _abrir_avancado(self):
        JanelaAvancado(self.root)


class JanelaAvancado(tk.Toplevel):
    """Janela separada com os scripts opcionais do repositorio original."""

    def __init__(self, master):
        super().__init__(master)
        self.title("Ferramentas avancadas (opcionais)")
        self.geometry("760x560")

        nb = ttk.Notebook(self)
        nb.pack(fill="both", expand=True, padx=10, pady=10)

        aba1 = ttk.Frame(nb, padding=12)
        aba2 = ttk.Frame(nb, padding=12)
        aba3 = ttk.Frame(nb, padding=12)
        nb.add(aba1, text="Gerar apendices (LDO/LOA)")
        nb.add(aba2, text="Validacao / Kappa (LDO/LOA)")
        nb.add(aba3, text="Reprocessamento (Instrumentos)")

        self._montar_apendices(aba1)
        self._montar_validacao(aba2)
        self._montar_reprocess(aba3)

    # -- helpers genericos --
    def _linha_arquivo(self, parent, label, var, salvar=False, tipos=(("Planilha Excel", "*.xlsx"),), defaultext=None,
                        initialfile=None):
        f = ttk.Frame(parent)
        f.pack(fill="x", pady=4)
        ttk.Label(f, text=label, width=30, anchor="w").pack(side="left")
        ttk.Entry(f, textvariable=var).pack(side="left", fill="x", expand=True, padx=6)

        def escolher():
            if salvar:
                caminho = filedialog.asksaveasfilename(filetypes=tipos, defaultextension=defaultext,
                                                         initialfile=initialfile)
            else:
                caminho = filedialog.askopenfilename(filetypes=tipos)
            if caminho:
                var.set(caminho)

        ttk.Button(f, text="Procurar...", command=escolher).pack(side="left")

    def _rodar_com_log(self, cmd, log_widget, cwd=None, env=None, ao_terminar=None):
        def escrever(linha):
            log_widget.config(state="normal")
            log_widget.insert("end", linha)
            log_widget.see("end")
            log_widget.config(state="disabled")

        log_widget.config(state="normal")
        log_widget.delete("1.0", "end")
        log_widget.config(state="disabled")
        escrever("$ " + " ".join(str(x) for x in cmd) + "\n")

        def alvo():
            try:
                proc = subprocess.Popen(cmd, cwd=cwd, env=env, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                                         text=True, encoding="utf-8", errors="replace", bufsize=1)
                for linha in proc.stdout:
                    self.after(0, escrever, linha)
                proc.wait()
                ok = proc.returncode == 0
            except Exception as e:
                self.after(0, escrever, f"\nERRO: {e}\n")
                ok = False
            if ao_terminar:
                self.after(0, lambda: ao_terminar(ok))

        threading.Thread(target=alvo, daemon=True).start()

    # -- aba 1: gerar apendices --
    def _montar_apendices(self, aba):
        ttk.Label(aba, text="Gera um .docx com o vocabulario e o livro de codigo a partir da planilha de "
                             "parametros, com carimbo SHA-256 (gerar_apendices.py).",
                  wraplength=700, justify="left").pack(anchor="w", pady=(0, 10))
        var_params = tk.StringVar(value=str(DEFAULT_PARAMETROS) if DEFAULT_PARAMETROS.exists() else "")
        var_saida = tk.StringVar(value="APENDICES_gerados.docx")
        self._linha_arquivo(aba, "Planilha de parametros:", var_params)
        self._linha_arquivo(aba, "Arquivo .docx de saida:", var_saida, salvar=True,
                             tipos=(("Documento Word", "*.docx"),), defaultext=".docx",
                             initialfile="APENDICES_gerados.docx")
        log = scrolledtext.ScrolledText(aba, height=16, state="disabled", font=("Consolas", 9))

        def rodar():
            if not os.path.isfile(var_params.get()):
                messagebox.showerror("Erro", "Escolha uma planilha de parametros valida.")
                return
            faltando = modulos_faltando(["docx", "openpyxl"])
            if faltando:
                messagebox.showerror("Faltam pacotes", f"Instale com:\n{comando_pip(faltando)}")
                return
            self._rodar_com_log(
                [sys.executable, str(SCRIPTS_LDO / "gerar_apendices.py"), var_params.get(), var_saida.get()], log)

        ttk.Button(aba, text="Gerar", command=rodar).pack(anchor="w", pady=8)
        log.pack(fill="both", expand=True)

    # -- aba 2: validacao / kappa --
    def _montar_validacao(self, aba):
        ttk.Label(aba, text="Amostrar: sorteia uma amostra dos resultados para codificacao manual por dois "
                             "avaliadores. Avaliar: depois de preencher as colunas rotulo_humano_1/2 na amostra, "
                             "calcula Kappa de Cohen e F1 (Harness_Validacao.py, semente fixa = 42).",
                  wraplength=700, justify="left").pack(anchor="w", pady=(0, 10))

        sub = ttk.Notebook(aba)
        sub.pack(fill="both", expand=True)
        t1 = ttk.Frame(sub, padding=8)
        t2 = ttk.Frame(sub, padding=8)
        sub.add(t1, text="Amostrar")
        sub.add(t2, text="Avaliar")

        var_res = tk.StringVar()
        var_amostra = tk.StringVar(value="amostra_para_codificar.xlsx")
        self._linha_arquivo(t1, "Resultados_LDO_LOA.xlsx:", var_res)
        self._linha_arquivo(t1, "Salvar amostra em:", var_amostra, salvar=True, defaultext=".xlsx",
                             initialfile="amostra_para_codificar.xlsx")
        log1 = scrolledtext.ScrolledText(t1, height=10, state="disabled", font=("Consolas", 9))

        def rodar_amostrar():
            faltando = modulos_faltando(["pandas", "openpyxl"])
            if faltando:
                messagebox.showerror("Faltam pacotes", f"Instale com:\n{comando_pip(faltando)}")
                return
            self._rodar_com_log(
                [sys.executable, str(SCRIPTS_LDO / "Harness_Validacao.py"), "amostrar",
                 var_res.get(), var_amostra.get()], log1)

        ttk.Button(t1, text="Amostrar", command=rodar_amostrar).pack(anchor="w", pady=6)
        log1.pack(fill="both", expand=True)

        var_preenchida = tk.StringVar()
        var_relatorio = tk.StringVar(value="relatorio_validacao.xlsx")
        self._linha_arquivo(t2, "Amostra preenchida:", var_preenchida)
        self._linha_arquivo(t2, "Salvar relatorio em:", var_relatorio, salvar=True, defaultext=".xlsx",
                             initialfile="relatorio_validacao.xlsx")
        log2 = scrolledtext.ScrolledText(t2, height=10, state="disabled", font=("Consolas", 9))

        def rodar_avaliar():
            faltando = modulos_faltando(["pandas", "sklearn"])
            if faltando:
                messagebox.showerror("Faltam pacotes", f"Instale com:\n{comando_pip(faltando)}")
                return
            self._rodar_com_log(
                [sys.executable, str(SCRIPTS_LDO / "Harness_Validacao.py"), "avaliar",
                 var_preenchida.get(), var_relatorio.get()], log2)

        ttk.Button(t2, text="Avaliar", command=rodar_avaliar).pack(anchor="w", pady=6)
        log2.pack(fill="both", expand=True)

    # -- aba 3: reprocessamento --
    def _montar_reprocess(self, aba):
        ttk.Label(aba,
                  text="Reprocessamento pos-parecer (reprocess.py): exclui documentos contaminados, aplica "
                       "falso-positivo/cadastro-fiscal-fraco e recalcula o Grau apenas com ocorrencias "
                       "aproveitaveis em instrumentos vigentes. PRECISA de um arquivo anexo_params.json que "
                       "nao foi incluido no pacote de codigo original (ver Readme_Instrumentos.md) e de uma "
                       "pasta de trabalho que ja tenha resultados.json, ocorrencias.json e analise_final.json "
                       "(gerados pelo passo principal da analise de Instrumentos).",
                  wraplength=700, justify="left").pack(anchor="w", pady=(0, 10))

        var_pasta = tk.StringVar()
        var_anexo = tk.StringVar()
        var_params = tk.StringVar(value=str(DEFAULT_PARAMETROS) if DEFAULT_PARAMETROS.exists() else "")

        f1 = ttk.Frame(aba)
        f1.pack(fill="x", pady=4)
        ttk.Label(f1, text="Pasta de trabalho:", width=30, anchor="w").pack(side="left")
        ttk.Entry(f1, textvariable=var_pasta).pack(side="left", fill="x", expand=True, padx=6)
        ttk.Button(f1, text="Procurar...", command=lambda: var_pasta.set(
            filedialog.askdirectory() or var_pasta.get())).pack(side="left")

        self._linha_arquivo(aba, "anexo_params.json:", var_anexo, tipos=(("JSON", "*.json"),))
        self._linha_arquivo(aba, "Planilha de parametros:", var_params)

        log = scrolledtext.ScrolledText(aba, height=14, state="disabled", font=("Consolas", 9))

        def rodar():
            pasta = var_pasta.get().strip()
            if not pasta or not os.path.isdir(pasta):
                messagebox.showerror("Erro", "Escolha uma pasta de trabalho valida.")
                return
            for nome in ("resultados.json", "ocorrencias.json", "analise_final.json"):
                if not os.path.isfile(os.path.join(pasta, nome)):
                    messagebox.showerror(
                        "Arquivo faltando",
                        f"Nao encontrei '{nome}' na pasta de trabalho. Rode primeiro a analise principal de "
                        "Instrumentos (Passo 3-4 do fluxo padrao) nessa mesma pasta.")
                    return
            if not var_anexo.get().strip() or not os.path.isfile(var_anexo.get()):
                messagebox.showerror("Erro", "Escolha um arquivo anexo_params.json valido.")
                return
            if not var_params.get().strip() or not os.path.isfile(var_params.get()):
                messagebox.showerror("Erro", "Escolha uma planilha de parametros valida.")
                return
            destino_anexo = os.path.join(pasta, "anexo_params.json")
            if os.path.abspath(var_anexo.get()) != os.path.abspath(destino_anexo):
                shutil.copyfile(var_anexo.get(), destino_anexo)
            env = os.environ.copy()
            env["PARAMETROS_XLSX"] = os.path.abspath(var_params.get())
            self._rodar_com_log([sys.executable, str(SCRIPTS_INSTR / "reprocess.py")], log, cwd=pasta, env=env)

        ttk.Button(aba, text="Rodar reprocessamento", command=rodar).pack(anchor="w", pady=8)
        log.pack(fill="both", expand=True)


def main():
    if openpyxl is None:
        # ainda assim abrimos a janela, mas avisamos: a leitura dos parametros gerais da LDO/LOA precisa dele.
        pass
    root = tk.Tk()
    try:
        style = ttk.Style()
        if "vista" in style.theme_names():
            style.theme_use("vista")
    except Exception:
        pass
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
